#!/usr/bin/env python3
"""
report_product_rates.py — Generate Products rate/MRP/selling-price xlsx report.

Output: Reports/product_rates_<date>.xlsx

Columns:
  short_name | volume_ml | carton_size | rate_per_cb | aroed_per_bottle | MRP | selling_price
"""

import os
from datetime import date
from decimal import Decimal, ROUND_HALF_UP

import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONN = dict(host='195.201.119.186', port=5432, dbname='imfl_app',
            user='postgres', password='P@ssword4postgres')

HEADERS = [
    'Short Name', 'Volume (ml)', 'Carton Size',
    'Rate Per CB', 'AROED Per Bottle', 'MRP', 'Selling Price',
]

COL_WIDTHS = [36, 13, 13, 14, 18, 12, 14]


def fetch_data(cur):
    cur.execute("""
        SELECT
            p.short_name,
            p.volume_ml,
            p.carton_size,
            pr.rate_per_cb,
            pr.aroed_per_bottle,
            csp.selling_price
        FROM "Product" p
        LEFT JOIN LATERAL (
            SELECT rate_per_cb, aroed_per_bottle
            FROM "ProductRate"
            WHERE product_id = p.id
            ORDER BY effective_date DESC, id DESC
            LIMIT 1
        ) pr ON true
        LEFT JOIN LATERAL (
            SELECT selling_price
            FROM "CounterSellingPrice"
            WHERE product_id = p.id
              AND effective_to IS NULL
            ORDER BY effective_from DESC
            LIMIT 1
        ) csp ON true
        WHERE p.is_active = true
        ORDER BY p.short_name
    """)
    return cur.fetchall()


def calc_mrp(rate_per_cb, aroed_per_bottle, carton_size):
    if rate_per_cb is None or carton_size is None or carton_size == 0:
        return None
    aroed = aroed_per_bottle if aroed_per_bottle is not None else Decimal('0')
    mrp = (Decimal(str(rate_per_cb)) / carton_size + Decimal(str(aroed))) * Decimal('1.10')
    return mrp.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def style_header_cell(cell):
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', fgColor='2E4057')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        bottom=Side(style='medium', color='FFFFFF'),
        right=Side(style='thin', color='CCCCCC'),
    )


def style_data_cell(cell, row_idx, is_mrp=False, is_number=False):
    bg = 'F2F5F8' if row_idx % 2 == 0 else 'FFFFFF'
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.border = Border(
        bottom=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
    )
    if is_mrp:
        cell.font = Font(bold=True, color='1A6B3C')
        cell.alignment = Alignment(horizontal='right')
    elif is_number:
        cell.alignment = Alignment(horizontal='right')
    else:
        cell.alignment = Alignment(horizontal='left')


def generate(out_path):
    conn = psycopg2.connect(**CONN)
    cur = conn.cursor()
    rows = fetch_data(cur)
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product Rates'

    # Title row
    ws.merge_cells('A1:G1')
    title = ws['A1']
    title.value = f'Product Rate & MRP Report  —  {date.today().strftime("%d %b %Y")}'
    title.font = Font(bold=True, size=13, color='2E4057')
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.fill = PatternFill('solid', fgColor='D9E4F0')
    ws.row_dimensions[1].height = 28

    # Header row
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=2, column=col, value=header)
        style_header_cell(cell)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 30

    # Freeze panes below header
    ws.freeze_panes = 'A3'

    # Data rows
    NUMBER_COLS = {2, 3, 4, 5, 6, 7}  # 1-based
    MRP_COL = 6

    for r_idx, row in enumerate(rows, 1):
        short_name, volume_ml, carton_size, rate_per_cb, aroed_per_bottle, selling_price = row
        mrp = calc_mrp(rate_per_cb, aroed_per_bottle, carton_size)

        values = [
            short_name,
            volume_ml,
            carton_size,
            float(rate_per_cb) if rate_per_cb is not None else None,
            float(aroed_per_bottle) if aroed_per_bottle is not None else None,
            float(mrp) if mrp is not None else None,
            float(selling_price) if selling_price is not None else None,
        ]

        excel_row = r_idx + 2
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            style_data_cell(
                cell, r_idx,
                is_mrp=(col == MRP_COL),
                is_number=(col in NUMBER_COLS),
            )
            if col in {4, 5, 6, 7} and val is not None:
                cell.number_format = '#,##0.00'
            if col in {2, 3} and val is not None:
                cell.number_format = '#,##0'

    # Summary row
    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value=f'Total Products: {len(rows)}').font = Font(bold=True, italic=True, color='555555')

    wb.save(out_path)
    print(f'Saved: {out_path}  ({len(rows)} products)')


if __name__ == '__main__':
    os.makedirs('Reports', exist_ok=True)
    out = f'Reports/product_rates_{date.today().strftime("%Y%m%d")}.xlsx'
    generate(out)
