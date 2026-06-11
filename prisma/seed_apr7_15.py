"""
Batch seed script — April 7–15, 2026 (Days 7–15).

Permit days: 7 (M7, 26 products, 6360 btls),
             11 (M11, 15 products, 3792 btls),
             15 (M15, 17 products, 5964 btls)
Day 14: shop closed — zeros seeded, no expenses/cash records.
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from decimal import Decimal
from datetime import date
from urllib.parse import urlparse, unquote
import os

raw_url = os.environ.get('DATABASE_URL',
    'postgresql://postgres:P%40ssword4postgres@195.201.119.186:5432/imfl_app')
raw_url = raw_url.split('?')[0]
u = urlparse(raw_url)
conn = psycopg2.connect(
    host=u.hostname, port=u.port or 5432,
    dbname=u.path.lstrip('/'), user=u.username, password=unquote(u.password),
)
conn.autocommit = False
cur = conn.cursor()

XLSX = '/home/prasadkabadi/ALK/Permits/KLS-I-APR-2026.xlsx'

def safe_int(v, default=0):
    try: return int(float(v)) if v is not None else default
    except: return default

def safe_dec(v, default='0'):
    try: return str(round(float(v), 2)) if v is not None else default
    except: return default

# ── References ─────────────────────────────────────────────────
cur.execute('SELECT id FROM "Shop" WHERE short_name = %s', ('KLS',))
shop_id = cur.fetchone()[0]
cur.execute('SELECT id, master_code FROM "Product"')
prod_id = {mc: pid for pid, mc in cur.fetchall()}
cur.execute('SELECT id, display_order FROM "Counter" WHERE shop_id = %s ORDER BY display_order', (shop_id,))
counter_id = {order: cid for cid, order in cur.fetchall()}
print(f'Shop={shop_id}, Products={len(prod_id)}, Counters={counter_id}')

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws_rm = wb['Rate Matrix']

def load_rate_matrix_mrp(col_index):
    mrp_map = {}
    for row in ws_rm.iter_rows(min_row=3, values_only=True):
        mc = row[0]
        if mc is None: continue
        mrp = row[col_index]
        if mrp is not None:
            mrp_map[int(float(mc))] = Decimal(str(mrp))
    return mrp_map

def counter_cols(row, base):
    return {
        'ob':     safe_int(row[base + 2]),
        'rec':    safe_int(row[base + 3]),
        'total':  safe_int(row[base + 4]),
        'sale':   safe_int(row[base + 5]),
        'cb':     safe_int(row[base + 6]),
        'amt':    safe_dec(row[base + 7]),
        'tips':   safe_dec(row[base + 8]),
        'totamt': safe_dec(row[base + 9]),
        'mrp':    safe_dec(row[base + 10]),
        'sp':     safe_dec(row[base + 11]),
    }

def godown_cols_permit(row):
    return {
        'ob':  safe_int(row[59]),
        'rec': safe_int(row[46]),   # permit day: shop-level receipt in col46
        'sc1': safe_int(row[62]),
        'sc2': safe_int(row[63]),
        'sc3': safe_int(row[64]),
        'cb':  safe_int(row[65]),
    }

def godown_cols_normal(row):
    return {
        'ob':  safe_int(row[59]),
        'rec': safe_int(row[60]),
        'sc1': safe_int(row[62]),
        'sc2': safe_int(row[63]),
        'sc3': safe_int(row[64]),
        'cb':  safe_int(row[65]),
    }

# ── Per-day config ─────────────────────────────────────────────
# (cnum, staff, liq, beer, tips, grand, gpay, exp, coll, drawer, tips_cash, tally)
DAY_DATA = {
    7: {
        'date': date(2026, 4, 7),
        'permit_mrp_col': 21,   # M7
        'summaries': [
            (1, 'Ramesh', '34387', '4245', '753', '38632', '5865', '6085', '30150', '3150', '753',  '3'),
            (2, None,     '16890', '1910', '415', '18800', '4095', '4235', '14415',  '565', '415',  '0'),
            (3, None,     '33336', '2370', '724', '35706', '6890', '6980', '27220', '2230', '724',  '4'),
        ],
        'expenses': [
            (1, 'TEA',        40,    False),
            (1, 'BREAKAGE',   130,   False),
            (1, 'GOOGLE_PAY', 5865,  True),
            (1, 'BHATTA',     50,    False),
            (2, 'TEA',        40,    False),
            (2, 'GOOGLE_PAY', 4095,  True),
            (2, 'BHATTA',     100,   False),
            (3, 'TEA',        40,    False),
            (3, 'GOOGLE_PAY', 6890,  True),
            (3, 'BHATTA',     50,    False),
        ],
        'cash': [
            # C1 collection
            (1, 'COLLECTION', 'NOTE', 500, 33),
            (1, 'COLLECTION', 'NOTE', 200, 12),
            (1, 'COLLECTION', 'NOTE', 100, 80),
            (1, 'COLLECTION', 'NOTE',  50, 50),
            # C2 collection
            (2, 'COLLECTION', 'NOTE', 500, 19),
            (2, 'COLLECTION', 'NOTE', 200,  5),
            (2, 'COLLECTION', 'NOTE', 100, 20),
            (2, 'COLLECTION', 'NOTE',  50, 30),
            # C3 collection
            (3, 'COLLECTION', 'NOTE', 500, 27),
            (3, 'COLLECTION', 'NOTE', 200, 45),
            (3, 'COLLECTION', 'NOTE', 100, 20),
            (3, 'COLLECTION', 'NOTE',  50, 40),
            # Drawer
            (1, 'DRAWER_CASH', 'COIN', 1, 3150),
            (2, 'DRAWER_CASH', 'COIN', 1,  565),
            (3, 'DRAWER_CASH', 'COIN', 1, 2230),
            # Tips
            (1, 'TIPS_CASH', 'COIN', 1, 753),
            (2, 'TIPS_CASH', 'COIN', 1, 415),
            (3, 'TIPS_CASH', 'COIN', 1, 724),
        ],
    },
    8: {
        'date': date(2026, 4, 8),
        'permit_mrp_col': None,
        'summaries': [
            (1, 'Ramesh', '32410', '1820', '715', '34230', '5495', '10585', '21520', '2840', '715', '-5'),
            (2, None,     '19822', '2620', '483', '22442', '4500',  '4590', '15085', '3250', '483', '-2'),
            (3, None,     '48817', '4995','1278', '53812', '9585',  '9675', '45275',  '140','1278',  '3'),
        ],
        'expenses': [
            (1, 'TEA',        40,    False),
            (1, 'PAK',        5000,  False),
            (1, 'GOOGLE_PAY', 5495,  True),
            (1, 'BHATTA',     50,    False),
            (2, 'TEA',        40,    False),
            (2, 'GOOGLE_PAY', 4500,  True),
            (2, 'BHATTA',     50,    False),
            (3, 'TEA',        40,    False),
            (3, 'GOOGLE_PAY', 9585,  True),
            (3, 'BHATTA',     50,    False),
        ],
        'cash': [
            (1, 'COLLECTION', 'NOTE', 500, 16),
            (1, 'COLLECTION', 'NOTE', 200, 19),
            (1, 'COLLECTION', 'NOTE', 100, 80),
            (1, 'COLLECTION', 'NOTE',  50, 20),
            (2, 'COLLECTION', 'NOTE', 500,  8),
            (2, 'COLLECTION', 'NOTE', 200, 18),
            (2, 'COLLECTION', 'NOTE', 100, 60),
            (2, 'COLLECTION', 'NOTE',  50, 20),
            (3, 'COLLECTION', 'NOTE', 500, 43),
            (3, 'COLLECTION', 'NOTE', 200, 70),
            (3, 'COLLECTION', 'NOTE', 100, 60),
            (3, 'COLLECTION', 'NOTE',  50, 40),
            (3, 'COLLECTION', 'NOTE',  20, 25),
            (1, 'DRAWER_CASH', 'COIN', 1, 2840),
            (2, 'DRAWER_CASH', 'COIN', 1, 3250),
            (3, 'DRAWER_CASH', 'COIN', 1,  140),
            (1, 'TIPS_CASH', 'COIN', 1,  715),
            (2, 'TIPS_CASH', 'COIN', 1,  483),
            (3, 'TIPS_CASH', 'COIN', 1, 1278),
        ],
    },
    9: {
        'date': date(2026, 4, 9),
        'permit_mrp_col': None,
        'summaries': [
            (1, 'Ramesh', '25759', '2430', '596', '28189', '6610', '7010', '18495', '3280', '596', '1'),
            (2, None,     '25470', '2160', '515', '27630', '4285', '4375', '20510', '3260', '515', '5'),
            (3, None,     '36662', '1495', '693', '38157', '4155', '4295', '33683',  '872', '683', '0'),
        ],
        'expenses': [
            (1, 'TEA',        40,    False),
            (1, 'PAK',        310,   False),
            (1, 'GOOGLE_PAY', 6610,  True),
            (1, 'BHATTA',     50,    False),
            (2, 'TEA',        40,    False),
            (2, 'GOOGLE_PAY', 4285,  True),
            (2, 'BHATTA',     50,    False),
            (3, 'TEA',        40,    False),
            (3, 'GOOGLE_PAY', 4155,  True),
            (3, 'BHATTA',     100,   False),
        ],
        'cash': [
            (1, 'COLLECTION', 'NOTE', 500, 22),
            (1, 'COLLECTION', 'NOTE', 200,  7),
            (1, 'COLLECTION', 'NOTE', 100, 35),
            (1, 'COLLECTION', 'NOTE',  50, 40),
            (2, 'COLLECTION', 'NOTE', 500, 11),
            (2, 'COLLECTION', 'NOTE', 200, 27),
            (2, 'COLLECTION', 'NOTE', 100, 60),
            (2, 'COLLECTION', 'NOTE',  50, 38),
            (2, 'COLLECTION', 'NOTE',  20, 60),
            (3, 'COLLECTION', 'NOTE', 500, 43),
            (3, 'COLLECTION', 'NOTE', 200, 21),
            (3, 'COLLECTION', 'NOTE', 100, 55),
            (3, 'COLLECTION', 'NOTE',  50, 36),
            (1, 'DRAWER_CASH', 'COIN', 1, 3280),
            (2, 'DRAWER_CASH', 'COIN', 1, 3260),
            (3, 'DRAWER_CASH', 'COIN', 1,  872),
            (1, 'TIPS_CASH', 'COIN', 1, 596),
            (2, 'TIPS_CASH', 'COIN', 1, 515),
            (3, 'TIPS_CASH', 'COIN', 1, 683),
        ],
    },
    10: {
        'date': date(2026, 4, 10),
        'permit_mrp_col': None,
        'summaries': [
            (1, 'Ramesh', '24485', '1740', '530', '26225', '5335', '5575', '18030', '3150', '530',  '0'),
            (2, None,     '25841', '2795', '529', '28636', '5990', '6080', '18785', '4300', '529', '-6'),
            (3, None,     '36129', '3535', '711', '39664', '5510', '5600', '32711', '2064', '711',  '0'),
        ],
        'expenses': [
            (1, 'TEA',        40,    False),
            (1, 'POOJA',      100,   False),
            (1, 'GOOGLE_PAY', 5335,  True),
            (1, 'BHATTA',     100,   False),
            (2, 'TEA',        40,    False),
            (2, 'GOOGLE_PAY', 5990,  True),
            (2, 'BHATTA',     50,    False),
            (3, 'TEA',        40,    False),
            (3, 'GOOGLE_PAY', 5510,  True),
            (3, 'BHATTA',     50,    False),
        ],
        'cash': [
            (1, 'COLLECTION', 'NOTE', 500, 22),
            (1, 'COLLECTION', 'NOTE', 200, 15),
            (1, 'COLLECTION', 'NOTE', 100, 20),
            (1, 'COLLECTION', 'NOTE',  50, 20),
            (1, 'COLLECTION', 'NOTE',  20, 25),
            (2, 'COLLECTION', 'NOTE', 500, 12),
            (2, 'COLLECTION', 'NOTE', 200, 15),
            (2, 'COLLECTION', 'NOTE', 100, 55),
            (2, 'COLLECTION', 'NOTE',  50, 50),
            (2, 'COLLECTION', 'NOTE',  20, 50),
            (2, 'COLLECTION', 'NOTE',  10, 25),
            (3, 'COLLECTION', 'NOTE', 500, 37),
            (3, 'COLLECTION', 'NOTE', 200, 19),
            (3, 'COLLECTION', 'NOTE', 100, 83),
            (3, 'COLLECTION', 'NOTE',  50, 28),
            (1, 'DRAWER_CASH', 'COIN', 1, 3150),
            (2, 'DRAWER_CASH', 'COIN', 1, 4300),
            (3, 'DRAWER_CASH', 'COIN', 1, 2064),
            (1, 'TIPS_CASH', 'COIN', 1, 530),
            (2, 'TIPS_CASH', 'COIN', 1, 529),
            (3, 'TIPS_CASH', 'COIN', 1, 711),
        ],
    },
    11: {
        'date': date(2026, 4, 11),
        'permit_mrp_col': 29,   # M11
        'summaries': [
            (1, 'Ramesh', '45857', '2320',  '898', '48177', '8560', '10700', '36395', '1980',  '898', '3'),
            (2, None,     '16383', '3490',  '437', '19873', '1945',  '4075', '11935', '4300',  '437', '2'),
            (3, None,     '38059', '5515',  '776', '43574', '5365',  '8175', '32776', '3399',  '776', '0'),
        ],
        'expenses': [
            (1, 'TEA',          40,    False),
            (1, 'PERMIT_RENT',  2050,  False),
            (1, 'GOOGLE_PAY',   8560,  True),
            (1, 'BHATTA',       50,    False),
            (2, 'TEA',          40,    False),
            (2, 'BREAKAGE',     170,   False),
            (2, 'GOOGLE_PAY',   1945,  True),
            (2, 'BHATTA',       50,    False),
            (2, 'PAK',          1870,  False),
            (3, 'TEA',          40,    False),
            (3, 'GOOGLE_PAY',   5365,  True),
            (3, 'BHATTA',       50,    False),
            (3, 'PAK',          2720,  False),
        ],
        'cash': [
            (1, 'COLLECTION', 'NOTE', 500, 62),
            (1, 'COLLECTION', 'NOTE', 200, 10),
            (1, 'COLLECTION', 'NOTE', 100, 10),
            (1, 'COLLECTION', 'NOTE',  50, 30),
            (2, 'COLLECTION', 'NOTE', 500,  7),
            (2, 'COLLECTION', 'NOTE', 200,  6),
            (2, 'COLLECTION', 'NOTE', 100, 50),
            (2, 'COLLECTION', 'NOTE',  50, 26),
            (2, 'COLLECTION', 'NOTE',  20, 25),
            (3, 'COLLECTION', 'NOTE', 500, 45),
            (3, 'COLLECTION', 'NOTE', 200, 12),
            (3, 'COLLECTION', 'NOTE', 100, 52),
            (3, 'COLLECTION', 'NOTE',  50, 22),
            (3, 'COLLECTION', 'NOTE',  20, 40),
            (1, 'DRAWER_CASH', 'COIN', 1, 1980),
            (2, 'DRAWER_CASH', 'COIN', 1, 4300),
            (3, 'DRAWER_CASH', 'COIN', 1, 3399),
            (1, 'TIPS_CASH', 'COIN', 1,  898),
            (2, 'TIPS_CASH', 'COIN', 1,  437),
            (3, 'TIPS_CASH', 'COIN', 1,  776),
        ],
    },
    12: {
        'date': date(2026, 4, 12),
        'permit_mrp_col': None,
        'summaries': [
            (1, 'Ramesh', '49729', '13850', '1756', '63579', '10024', '10114', '53761', '1460', '1756', '-5'),
            (2, None,     '22293',  '5180',  '697', '27473',  '3415',  '3505', '19795', '4870',  '697',  '2'),
            (3, None,     '45239',  '8290', '1156', '53529', '12195', '12285', '42156',  '244', '1156',  '0'),
        ],
        'expenses': [
            (1, 'TEA',        40,    False),
            (1, 'GOOGLE_PAY', 10024, True),
            (1, 'BHATTA',     50,    False),
            (2, 'TEA',        40,    False),
            (2, 'GOOGLE_PAY', 3415,  True),
            (2, 'BHATTA',     50,    False),
            (3, 'TEA',        40,    False),
            (3, 'GOOGLE_PAY', 12195, True),
            (3, 'BHATTA',     50,    False),
        ],
        'cash': [
            (1, 'COLLECTION', 'NOTE', 500, 76),
            (1, 'COLLECTION', 'NOTE', 200, 45),
            (1, 'COLLECTION', 'NOTE', 100, 40),
            (1, 'COLLECTION', 'NOTE',  50, 20),
            (2, 'COLLECTION', 'NOTE', 500, 22),
            (2, 'COLLECTION', 'NOTE', 200, 15),
            (2, 'COLLECTION', 'NOTE', 100, 30),
            (2, 'COLLECTION', 'NOTE',  50, 30),
            (2, 'COLLECTION', 'NOTE',  20, 30),
            (3, 'COLLECTION', 'NOTE', 500, 62),
            (3, 'COLLECTION', 'NOTE', 200, 22),
            (3, 'COLLECTION', 'NOTE', 100, 41),
            (3, 'COLLECTION', 'NOTE',  50, 20),
            (3, 'COLLECTION', 'NOTE',  20, 25),
            (1, 'DRAWER_CASH', 'COIN', 1, 1460),
            (2, 'DRAWER_CASH', 'COIN', 1, 4870),
            (3, 'DRAWER_CASH', 'COIN', 1,  244),
            (1, 'TIPS_CASH', 'COIN', 1, 1756),
            (2, 'TIPS_CASH', 'COIN', 1,  697),
            (3, 'TIPS_CASH', 'COIN', 1, 1156),
        ],
    },
    13: {
        'date': date(2026, 4, 13),
        'permit_mrp_col': None,
        'summaries': [
            (1, 'Ramesh', '62100', '8160', '1435', '70260', '13320', '13510', '53435', '4750', '1435', '0'),
            (2, None,     '30884', '4400',  '541', '35284',  '5115',  '5205', '25040', '5580',  '541', '1'),
            (3, None,     '53585', '7405',  '940', '60990',  '8385',  '8595', '51440', '1895',  '940', '0'),
        ],
        'expenses': [
            (1, 'TEA',        40,    False),
            (1, 'PAK',        100,   False),
            (1, 'GOOGLE_PAY', 13320, True),
            (1, 'BHATTA',     50,    False),
            (2, 'TEA',        40,    False),
            (2, 'GOOGLE_PAY', 5115,  True),
            (2, 'BHATTA',     50,    False),
            (3, 'TEA',        40,    False),
            (3, 'BREAKAGE',   120,   False),
            (3, 'GOOGLE_PAY', 8385,  True),
            (3, 'BHATTA',     50,    False),
        ],
        'cash': [
            (1, 'COLLECTION', 'NOTE', 500, 68),
            (1, 'COLLECTION', 'NOTE', 200, 50),
            (1, 'COLLECTION', 'NOTE', 100, 70),
            (1, 'COLLECTION', 'NOTE',  50, 20),
            (2, 'COLLECTION', 'NOTE', 500, 27),
            (2, 'COLLECTION', 'NOTE', 200, 15),
            (2, 'COLLECTION', 'NOTE', 100, 65),
            (2, 'COLLECTION', 'NOTE',  50, 30),
            (3, 'COLLECTION', 'NOTE', 500, 75),
            (3, 'COLLECTION', 'NOTE', 200, 22),
            (3, 'COLLECTION', 'NOTE', 100, 76),
            (3, 'COLLECTION', 'NOTE',  50, 20),
            (1, 'DRAWER_CASH', 'COIN', 1, 4750),
            (2, 'DRAWER_CASH', 'COIN', 1, 5580),
            (3, 'DRAWER_CASH', 'COIN', 1, 1895),
            (1, 'TIPS_CASH', 'COIN', 1, 1435),
            (2, 'TIPS_CASH', 'COIN', 1,  541),
            (3, 'TIPS_CASH', 'COIN', 1,  940),
        ],
    },
    14: {
        'date': date(2026, 4, 14),
        'permit_mrp_col': None,
        'closed': True,   # shop closed — all zeros
        'summaries': [
            (1, 'Ramesh', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0'),
            (2, None,     '0', '0', '0', '0', '0', '0', '0', '0', '0', '0'),
            (3, None,     '0', '0', '0', '0', '0', '0', '0', '0', '0', '0'),
        ],
        'expenses': [],
        'cash': [],
    },
    15: {
        'date': date(2026, 4, 15),
        'permit_mrp_col': 37,   # M15
        'summaries': [
            (1, 'Ramesh', '36510', '3425',  '960', '39935',  '7595',  '7785', '30460', '2650',  '960', '0'),
            (2, None,     '17950', '2085',  '365', '20035',  '4195',  '5665', '10965', '3770',  '365', '0'),
            (3, None,     '45060', '8515', '1130', '53575', '11235', '11325', '43130',  '250', '1130', '0'),
        ],
        'expenses': [
            (1, 'TEA',          40,    False),
            (1, 'PAK',          100,   False),
            (1, 'GOOGLE_PAY',   7595,  True),
            (1, 'BHATTA',       50,    False),
            (2, 'TEA',          40,    False),
            (2, 'PERMIT_RENT',  1360,  False),
            (2, 'PAK',          20,    False),
            (2, 'GOOGLE_PAY',   4195,  True),
            (2, 'BHATTA',       50,    False),
            (3, 'TEA',          40,    False),
            (3, 'GOOGLE_PAY',   11235, True),
            (3, 'BHATTA',       50,    False),
        ],
        'cash': [
            (1, 'COLLECTION', 'NOTE', 500, 54),
            (1, 'COLLECTION', 'NOTE', 100, 10),
            (1, 'COLLECTION', 'NOTE',  50, 30),
            (2, 'COLLECTION', 'NOTE', 500,  8),
            (2, 'COLLECTION', 'NOTE', 200,  2),
            (2, 'COLLECTION', 'NOTE', 100, 50),
            (2, 'COLLECTION', 'NOTE',  50, 10),
            (2, 'COLLECTION', 'NOTE',  20, 35),
            (3, 'COLLECTION', 'NOTE', 500, 57),
            (3, 'COLLECTION', 'NOTE', 200, 40),
            (3, 'COLLECTION', 'NOTE', 100, 45),
            (3, 'COLLECTION', 'NOTE',  50, 20),
            (1, 'DRAWER_CASH', 'COIN', 1, 2650),
            (2, 'DRAWER_CASH', 'COIN', 1, 3770),
            (3, 'DRAWER_CASH', 'COIN', 1,  250),
            (1, 'TIPS_CASH', 'COIN', 1,  960),
            (2, 'TIPS_CASH', 'COIN', 1,  365),
            (3, 'TIPS_CASH', 'COIN', 1, 1130),
        ],
    },
}

# ── Process each day ───────────────────────────────────────────
for day_num in range(7, 16):
    cfg = DAY_DATA[day_num]
    day_date = cfg['date']
    is_permit = cfg['permit_mrp_col'] is not None
    is_closed = cfg.get('closed', False)

    print(f'\n{"="*60}')
    print(f'Processing April {day_num}, 2026', end='')
    if is_permit: print(f' [PERMIT M{["","","","","","","","7","","","","11","","","","15"][day_num]}]', end='')
    if is_closed: print(' [CLOSED]', end='')
    print()
    print('='*60)

    ws = wb[str(day_num)]

    # Load MRP for permit day
    rate_mrp = {}
    if is_permit:
        rate_mrp = load_rate_matrix_mrp(cfg['permit_mrp_col'])

    # Parse product rows
    stock_c = {1: [], 2: [], 3: []}
    stock_gd = []
    permit_lots = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2 or i >= 420: continue
        mc = row[0]
        if mc is None or not isinstance(mc, (int, float)): continue
        mc = int(mc)
        pid = prod_id.get(mc)
        if not pid: continue

        stock_c[1].append((pid, counter_cols(row, 1)))
        stock_c[2].append((pid, counter_cols(row, 15)))
        stock_c[3].append((pid, counter_cols(row, 29)))

        if is_permit:
            gd = godown_cols_permit(row)
            stock_gd.append((pid, gd))
            permit_qty = safe_int(row[46])
            if permit_qty > 0:
                mrp = rate_mrp.get(mc)
                if mrp:
                    permit_lots.append((pid, permit_qty, mrp))
        else:
            stock_gd.append((pid, godown_cols_normal(row)))

    print(f'  Product rows: {len(stock_c[1])}', end='')
    if is_permit: print(f', permit lots: {len(permit_lots)} products, {sum(q for _,q,_ in permit_lots)} btls', end='')
    print()

    # DailyCounterStock
    for cnum in [1, 2, 3]:
        rows = []
        for pid, d in stock_c[cnum]:
            rows.append((counter_id[cnum], pid, day_date,
                         d['ob'], d['rec'], d['total'], d['sale'], d['cb'],
                         d['mrp'], d['sp'], d['amt'], d['tips'], d['totamt']))
        execute_values(cur, """
            INSERT INTO "DailyCounterStock"
                (counter_id, product_id, stock_date,
                 opening_balance_btls, received_from_godown_btls, total_btls,
                 sold_btls, closing_balance_btls,
                 mrp_per_bottle, selling_price_per_bottle,
                 sale_amount_mrp, tips_amount, total_sale_amount)
            SELECT v.cid, v.pid, v.dt::date,
                   v.ob::int, v.rec::int, v.total::int,
                   v.sale::int, v.cb::int,
                   v.mrp::numeric, v.sp::numeric,
                   v.amt::numeric, v.tips::numeric, v.totamt::numeric
            FROM (VALUES %s) AS v(cid,pid,dt,ob,rec,total,sale,cb,mrp,sp,amt,tips,totamt)
            ON CONFLICT DO NOTHING
        """, rows)
    print(f'  DailyCounterStock: {len(stock_c[1]) * 3} rows')

    # GodownStock
    gd_rows = [(shop_id, pid, day_date, gd['ob'], gd['rec'], gd['cb']) for pid, gd in stock_gd]
    execute_values(cur, """
        INSERT INTO "GodownStock" (shop_id, product_id, stock_date,
                                   opening_balance_btls, received_btls, closing_balance_btls)
        SELECT v.shop_id, v.pid, v.dt::date, v.ob::int, v.rec::int, v.cb::int
        FROM (VALUES %s) AS v(shop_id, pid, dt, ob, rec, cb)
        ON CONFLICT (shop_id, product_id, stock_date) DO NOTHING
    """, gd_rows)

    cur.execute('SELECT id, product_id FROM "GodownStock" WHERE shop_id = %s AND stock_date = %s',
                (shop_id, day_date))
    gd_stock_id = {row[1]: row[0] for row in cur.fetchall()}

    # GodownDistribution
    dist_rows = []
    for pid, gd in stock_gd:
        gsid = gd_stock_id.get(pid)
        if not gsid: continue
        for cnum, qty in [(1, gd['sc1']), (2, gd['sc2']), (3, gd['sc3'])]:
            if qty > 0:
                dist_rows.append((gsid, counter_id[cnum], qty, day_date))
    if dist_rows:
        execute_values(cur, """
            INSERT INTO "GodownDistribution" (godown_stock_id, counter_id, distributed_btls, distribution_date)
            SELECT v.gsid, v.cid, v.qty::int, v.dt::date
            FROM (VALUES %s) AS v(gsid, cid, qty, dt)
        """, dist_rows)
    print(f'  GodownStock: {len(gd_rows)} rows, GodownDist: {len(dist_rows)} rows')

    # StockLot (permit days only)
    if is_permit and permit_lots:
        cur.execute("""
            SELECT product_id FROM "StockLot"
            WHERE shop_id = %s AND lot_date = %s AND source = 'PERMIT'
        """, (shop_id, day_date))
        already_done = {r[0] for r in cur.fetchall()}
        lot_rows = [(shop_id, pid, mrp, qty, 0, day_date)
                    for pid, qty, mrp in permit_lots if pid not in already_done]
        if lot_rows:
            cur.executemany("""
                INSERT INTO "StockLot"
                    (shop_id, product_id, mrp_per_bottle, initial_qty, consumed_qty,
                     lot_date, source, permit_item_id)
                VALUES (%s, %s, %s, %s, %s, %s, 'PERMIT', NULL)
            """, lot_rows)
        print(f'  StockLot (PERMIT): {len(lot_rows)} rows')

    # DailyCounterSummary
    for cnum, staff, liq, beer, tips, grand, gpay, exp, coll, drawer, tips_cash, tally in cfg['summaries']:
        cur.execute("""
            INSERT INTO "DailyCounterSummary"
                (counter_id, summary_date, staff_name,
                 liquor_sale_amount, beer_sale_amount, total_tips, grand_total_by_sale,
                 google_pay_amount, expenses_total, collection_total,
                 drawer_cash_total, tips_cash_total, tally_difference)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (counter_id, summary_date) DO NOTHING
        """, (counter_id[cnum], day_date, staff,
              liq, beer, tips, grand, gpay, exp, coll, drawer, tips_cash, tally))

    # DailyExpense
    for cnum, category, amount, is_upi in cfg['expenses']:
        cur.execute("""
            INSERT INTO "DailyExpense" (counter_id, expense_date, category, amount, is_upi)
            VALUES (%s, %s, %s, %s, %s)
        """, (counter_id[cnum], day_date, category, str(amount), is_upi))

    # CashRecord
    for cnum, cat, dtype, dval, count in cfg['cash']:
        cur.execute("""
            INSERT INTO "CashRecord"
                (counter_id, record_date, cash_category, denomination_type,
                 denomination_value, count, total_amount)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (counter_id, record_date, cash_category, denomination_type, denomination_value)
            DO NOTHING
        """, (counter_id[cnum], day_date, cat, dtype, dval, count, dval * count))

    print(f'  Expenses: {len(cfg["expenses"])}, CashRecords: {len(cfg["cash"])}')

conn.commit()
cur.close()
conn.close()
print('\n✓ April 7–15, 2026 batch seed complete.')
