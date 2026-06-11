#!/usr/bin/env python3
"""
seed_apr16_30.py — Generic KSBCL daily seeder (reads any KLS-*.xlsx)
Inserts: GodownStock, DailyCounterStock, DailyCounterSummary,
         DailyExpense, CashRecord  (one sheet per day)

Usage:
  python3 seed_apr16_30.py                                    # Apr 16-30 defaults
  python3 seed_apr16_30.py --dry-run
  python3 seed_apr16_30.py --day 18
  python3 seed_apr16_30.py --xlsx /path/to/file.xlsx --month 5 --start 1 --end 15
"""

import argparse, os, sys, subprocess
from datetime import date
from decimal import Decimal

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── DB connection ─────────────────────────────────────────────────────────────
CONN = dict(host="195.201.119.186", port=5432,
            dbname="imfl_app", user="postgres", password="P@ssword4postgres")

XLSX = "/home/prasadkabadi/ALK/Permits/KLS-II-APR-2026.xlsx"
SHOP = "KLS"
YEAR = 2026

GENERATE_SCRIPT = os.path.join(os.path.dirname(__file__), "generate_daily_bills.py")

# ── Column layout (0-indexed, same template as KLS-I) ────────────────────────
# Counter bases: C1=1, C2=15, C3=29
# Within each counter: ob=base+2, rec=base+3, total=base+4,
#                      sale=base+5, cb=base+6, amt=base+7, tips=base+8,
#                      totamt=base+9, mrp=base+10, sp=base+11
# Godown: ob=59, rec=60, sc1=62, sc2=63, sc3=64, cb=65

CTR_BASES = {1: 1, 2: 15, 3: 29}

# Summary rows (0-indexed)
R_STAFF       = 423   # staff names: col[1]/[15]/[29]
R_LIQ         = 425   # liquor:  col[6]=sale, col[7]=tips  (C1); same +14/+28 for C2/C3
R_BEER        = 426   # beer:    col[6]=sale, col[7]=tips, col[11]=tally
R_GTOTAL_ROW  = 444   # col[8]=C1 collection grand, col[13]=C1 drawer,
                      # col[22]=C2 coll, col[27]=C2 drawer,
                      # col[36]=C3 coll, col[41]=C3 drawer
R_GPAY        = 439   # Google Pay row (same as EXPENSE_ROWS entry 439)
R_EXP_TOTAL   = 445   # expenses total: col[2]=C1, col[16]=C2, col[30]=C3
R_TIPS_CASH   = 460   # tips cash:      col[8]=C1, col[22]=C2, col[36]=C3

# Expense rows (0-indexed) → (category, is_upi)
EXPENSE_ROWS = [
    (433, 'TEA',                 False),
    (434, 'PERMIT_RENT',         False),
    (435, 'POOJA',               False),
    (436, 'BREAKAGE',            False),
    (437, 'OVER_CASH',           False),
    (438, 'ELECTRICITY_BILL',    False),
    (439, 'GOOGLE_PAY',          True ),
    (440, 'BHATTA',              False),
    (441, 'GOOGLE_PAY_NEGATIVE', False),
    (442, 'PAK',                 False),  # unlabeled row, holds PAK amount
    (444, 'OTHERS',              False),
]
# Expense amount columns: C1=col[2], C2=col[16], C3=col[30]
EXP_COLS = {1: 2, 2: 16, 3: 30}

# Collection denomination columns per counter (type, denom, count, total)
COLL_COLS = {1: (5, 6, 7, 8), 2: (19, 20, 21, 22), 3: (33, 34, 35, 36)}
# Drawer cash columns per counter (type, denom, count, total) — row R_GTOTAL_ROW
DRAW_TOTAL_COL = {1: 13, 2: 27, 3: 41}
COLL_TOTAL_COL = {1: 8,  2: 22, 3: 36}
# Tips cash grand total row R_TIPS_CASH: same columns as COLL_TOTAL_COL
TIPS_TOTAL_COL = {1: 8,  2: 22, 3: 36}

# Summary cols for liquor/beer per counter
LIQ_SALE_COL  = {1: 6,  2: 20, 3: 34}
LIQ_TIPS_COL  = {1: 7,  2: 21, 3: 35}
BEER_SALE_COL = {1: 6,  2: 20, 3: 34}
BEER_TIPS_COL = {1: 7,  2: 21, 3: 35}
TALLY_COL     = {1: 11, 2: 25, 3: 39}
GPAY_COL      = {1: 2,  2: 16, 3: 30}
STAFF_COL     = {1: 1,  2: 15, 3: 29}


def si(v, default=0):
    try: return int(float(v)) if v is not None else default
    except: return default

def sd(v, default='0'):
    try: return str(round(float(v), 2)) if v is not None else default
    except: return default

def sfloat(v, default=0.0):
    try: return float(v) if v is not None else default
    except: return default


def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    return rows


def extract_stock(rows):
    """
    Returns ({cnum: [(pid_mc, col_dict), ...]}, [(pid_mc, gd_dict), ...]).
    pid_mc is the master_code (int) — caller maps to actual product IDs.
    """
    stock_c = {1: [], 2: [], 3: []}
    stock_gd = []

    for i, row in enumerate(rows):
        if i < 2 or i >= 420:
            continue
        mc = row[0]
        if mc is None or not isinstance(mc, (int, float)):
            continue
        mc = int(mc)

        for cnum, base in CTR_BASES.items():
            stock_c[cnum].append((mc, {
                'ob':     si(row[base + 2]),
                'rec':    si(row[base + 3]),
                'total':  si(row[base + 4]),
                'sale':   si(row[base + 5]),
                'cb':     si(row[base + 6]),
                'amt':    sd(row[base + 7]),
                'tips':   sd(row[base + 8]),
                'totamt': sd(row[base + 9]),
                'mrp':    sd(row[base + 10]),
                'sp':     sd(row[base + 11]),
            }))

        stock_gd.append((mc, {
            'ob':  si(row[59]),
            'rec': si(row[60]),
            'sc1': si(row[62]),
            'sc2': si(row[63]),
            'sc3': si(row[64]),
            'cb':  si(row[65]),
        }))

    return stock_c, stock_gd


def extract_summary(rows):
    """Returns {cnum: summary_dict}."""
    result = {}
    for cnum in [1, 2, 3]:
        liq_sale = sfloat(rows[R_LIQ ][LIQ_SALE_COL [cnum]])
        liq_tips = sfloat(rows[R_LIQ ][LIQ_TIPS_COL [cnum]])
        beer_sale= sfloat(rows[R_BEER][BEER_SALE_COL[cnum]])
        beer_tips= sfloat(rows[R_BEER][BEER_TIPS_COL[cnum]])
        tally    = sfloat(rows[R_BEER][TALLY_COL    [cnum]])
        gpay     = sfloat(rows[R_GPAY][GPAY_COL[cnum]])  # R_GTOTAL_ROW-5 = row 439
        exp_tot  = sfloat(rows[R_EXP_TOTAL][EXP_COLS[cnum]])
        coll     = sfloat(rows[R_GTOTAL_ROW][COLL_TOTAL_COL[cnum]])
        drawer   = sfloat(rows[R_GTOTAL_ROW][DRAW_TOTAL_COL[cnum]])
        tips_c   = sfloat(rows[R_TIPS_CASH ][TIPS_TOTAL_COL[cnum]])
        staff    = rows[R_STAFF][STAFF_COL[cnum]]

        result[cnum] = {
            'staff':       staff,
            'liq_sale':    liq_sale,
            'beer_sale':   beer_sale,
            'total_tips':  liq_tips + beer_tips,
            'grand':       liq_sale + beer_sale,
            'gpay':        gpay,
            'exp_total':   exp_tot,
            'collection':  coll,
            'drawer':      drawer,
            'tips_cash':   tips_c,
            'tally':       tally,
        }
    return result


def extract_expenses(rows):
    """Returns {cnum: [(category, amount, is_upi), ...]} — only non-zero entries."""
    result = {1: [], 2: [], 3: []}
    for row_i, cat, is_upi in EXPENSE_ROWS:
        for cnum in [1, 2, 3]:
            amt = sfloat(rows[row_i][EXP_COLS[cnum]])
            if amt > 0:
                result[cnum].append((cat, amt, is_upi))
    return result


def extract_cash(rows):
    """
    Returns {cnum: [(cash_category, denom_type, denom_val, count, total), ...]}.
    COLLECTION: scanned from denomination rows 433-444.
    DRAWER_CASH and TIPS_CASH: single COIN entries.
    """
    result = {1: [], 2: [], 3: []}

    # Collection denominations (rows i=433 to i=443)
    for i in range(433, 444):
        row = rows[i]
        for cnum, (tc, dc, cc, totc) in COLL_COLS.items():
            dtype  = row[tc]
            dval   = sfloat(row[dc])
            count  = sfloat(row[cc])
            total  = sfloat(row[totc])
            dtype_db = 'NOTE' if dtype == 'Notes' else 'COIN' if dtype == 'Coins' else None
            if dtype_db and count > 0 and total > 0:
                result[cnum].append(('COLLECTION', dtype_db, int(dval), int(count), int(total)))

    # Drawer cash (single COIN entry)
    for cnum in [1, 2, 3]:
        d = int(sfloat(rows[R_GTOTAL_ROW][DRAW_TOTAL_COL[cnum]]))
        if d > 0:
            result[cnum].append(('DRAWER_CASH', 'COIN', 1, d, d))

    # Tips cash (single COIN entry)
    for cnum in [1, 2, 3]:
        t = int(sfloat(rows[R_TIPS_CASH][TIPS_TOTAL_COL[cnum]]))
        if t > 0:
            result[cnum].append(('TIPS_CASH', 'COIN', 1, t, t))

    return result


def seed_day(cur, shop_id, prod_id, counter_id, sale_date, rows, dry_run=False):
    stock_c, stock_gd = extract_stock(rows)
    summary            = extract_summary(rows)
    expenses           = extract_expenses(rows)
    cash               = extract_cash(rows)

    # Check if this is a sales day
    total_sold = sum(d['sale'] for cnum in [1,2,3] for _, d in stock_c[cnum])
    if total_sold == 0:
        print(f"  ⚠  No sales found — skipping {sale_date}")
        return False

    print(f"\n  ── {sale_date}  (total bottles sold: {total_sold}) ──")
    for cnum in [1,2,3]:
        s = summary[cnum]
        print(f"    C{cnum}: liq={s['liq_sale']:.0f}  beer={s['beer_sale']:.0f}  "
              f"tips={s['total_tips']:.0f}  gpay={s['gpay']:.0f}  "
              f"coll={s['collection']:.0f}  drawer={s['drawer']:.0f}  tally={s['tally']:.0f}")

    if dry_run:
        return True

    # ── DailyCounterStock ─────────────────────────────────────────
    for cnum in [1, 2, 3]:
        rows_db = []
        for mc, d in stock_c[cnum]:
            pid = prod_id.get(mc)
            if not pid:
                continue
            rows_db.append((counter_id[cnum], pid, sale_date,
                            d['ob'], d['rec'], d['total'], d['sale'], d['cb'],
                            d['mrp'], d['sp'], d['amt'], d['tips'], d['totamt']))
        execute_values(cur, """
            INSERT INTO "DailyCounterStock"
                (counter_id, product_id, stock_date,
                 opening_balance_btls, received_from_godown_btls, total_btls,
                 sold_btls, closing_balance_btls,
                 mrp_per_bottle, selling_price_per_bottle,
                 sale_amount_mrp, tips_amount, total_sale_amount)
            SELECT v.cid, v.pid, v.dt::date,
                   v.ob::int, v.rec::int, v.total::int,
                   v.sale::int, v.cb::int,
                   v.mrp::numeric, v.sp::numeric,
                   v.amt::numeric, v.tips::numeric, v.totamt::numeric
            FROM (VALUES %s) AS v(cid,pid,dt,ob,rec,total,sale,cb,mrp,sp,amt,tips,totamt)
            ON CONFLICT DO NOTHING
        """, rows_db)

    # ── GodownStock ───────────────────────────────────────────────
    gd_rows = []
    for mc, gd in stock_gd:
        pid = prod_id.get(mc)
        if pid:
            gd_rows.append((shop_id, pid, sale_date, gd['ob'], gd['rec'], gd['cb']))
    execute_values(cur, """
        INSERT INTO "GodownStock" (shop_id, product_id, stock_date,
                                   opening_balance_btls, received_btls, closing_balance_btls)
        SELECT v.shop_id, v.pid, v.dt::date, v.ob::int, v.rec::int, v.cb::int
        FROM (VALUES %s) AS v(shop_id, pid, dt, ob, rec, cb)
        ON CONFLICT (shop_id, product_id, stock_date) DO NOTHING
    """, gd_rows)

    cur.execute('SELECT id, product_id FROM "GodownStock" WHERE shop_id=%s AND stock_date=%s',
                (shop_id, sale_date))
    gd_stock_id = {r[1]: r[0] for r in cur.fetchall()}

    # ── GodownDistribution ────────────────────────────────────────
    dist_rows = []
    for mc, gd in stock_gd:
        pid = prod_id.get(mc)
        gsid = gd_stock_id.get(pid)
        if not gsid:
            continue
        for cnum, qty in [(1, gd['sc1']), (2, gd['sc2']), (3, gd['sc3'])]:
            if qty > 0:
                dist_rows.append((gsid, counter_id[cnum], qty, sale_date))
    if dist_rows:
        execute_values(cur, """
            INSERT INTO "GodownDistribution"
                (godown_stock_id, counter_id, distributed_btls, distribution_date)
            SELECT v.gsid, v.cid, v.qty::int, v.dt::date
            FROM (VALUES %s) AS v(gsid, cid, qty, dt)
        """, dist_rows)

    # ── DailyCounterSummary ───────────────────────────────────────
    for cnum in [1, 2, 3]:
        s = summary[cnum]
        cur.execute("""
            INSERT INTO "DailyCounterSummary"
                (counter_id, summary_date, staff_name,
                 liquor_sale_amount, beer_sale_amount, total_tips, grand_total_by_sale,
                 google_pay_amount, expenses_total, collection_total,
                 drawer_cash_total, tips_cash_total, tally_difference)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (counter_id, summary_date) DO NOTHING
        """, (counter_id[cnum], sale_date, s['staff'],
              str(s['liq_sale']), str(s['beer_sale']), str(s['total_tips']),
              str(s['grand']), str(s['gpay']), str(s['exp_total']),
              str(s['collection']), str(s['drawer']), str(s['tips_cash']),
              str(s['tally'])))

    # ── DailyExpense ──────────────────────────────────────────────
    for cnum in [1, 2, 3]:
        for cat, amt, is_upi in expenses[cnum]:
            cur.execute("""
                INSERT INTO "DailyExpense"
                    (counter_id, expense_date, category, amount, is_upi)
                VALUES (%s,%s,%s,%s,%s)
            """, (counter_id[cnum], sale_date, cat, str(amt), is_upi))

    # ── CashRecord ────────────────────────────────────────────────
    for cnum in [1, 2, 3]:
        for cat, dtype, dval, count, total in cash[cnum]:
            cur.execute("""
                INSERT INTO "CashRecord"
                    (counter_id, record_date, cash_category, denomination_type,
                     denomination_value, count, total_amount)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (counter_id, record_date, cash_category, denomination_type, denomination_value)
                DO NOTHING
            """, (counter_id[cnum], sale_date, cat, dtype, dval, count, total))

    print(f"    ✓ Seeded to DB.")
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--day',      type=int,   help='Single day number')
    ap.add_argument('--start',    type=int,   default=16, help='First day (default 16)')
    ap.add_argument('--end',      type=int,   default=30, help='Last day inclusive (default 30)')
    ap.add_argument('--month',    type=int,   default=4,  help='Month number (default 4=April)')
    ap.add_argument('--xlsx',     type=str,   default=XLSX, help='Path to Excel file')
    ap.add_argument('--dry-run',  action='store_true')
    ap.add_argument('--no-bills', action='store_true', help='Skip bill generation')
    args = ap.parse_args()

    xlsx_path = args.xlsx
    month     = args.month
    days = [args.day] if args.day else list(range(args.start, args.end + 1))

    conn = psycopg2.connect(**CONN)
    conn.autocommit = False
    cur  = conn.cursor()

    cur.execute('SELECT id FROM "Shop" WHERE short_name=%s', (SHOP,))
    shop_id = cur.fetchone()[0]

    cur.execute('SELECT id, master_code FROM "Product"')
    prod_id = {mc: pid for pid, mc in cur.fetchall()}

    cur.execute('SELECT id, display_order FROM "Counter" WHERE shop_id=%s ORDER BY display_order',
                (shop_id,))
    counter_id = {order: cid for cid, order in cur.fetchall()}

    print(f"Shop={shop_id}  Products={len(prod_id)}  Counters={counter_id}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    active_days = []
    for day in days:
        sheet_name = str(day)
        if sheet_name not in wb.sheetnames:
            print(f"\n  Sheet '{sheet_name}' not found — skipping.")
            continue

        sale_date = date(YEAR, month, day)
        print(f"\nProcessing sheet '{sheet_name}' → {sale_date}")
        rows = read_sheet(wb[sheet_name])

        seeded = seed_day(cur, shop_id, prod_id, counter_id, sale_date, rows,
                          dry_run=args.dry_run)
        if seeded and not args.dry_run:
            conn.commit()
            active_days.append(day)

    cur.close()
    conn.close()

    if not active_days:
        print("\nNo active days to generate bills for.")
        return

    if args.no_bills or args.dry_run:
        print(f"\nSeeded days: {active_days}")
        print("Run generate_daily_bills.py separately for each day.")
        return

    print(f"\n{'─'*50}")
    print("Generating bills for active days...")
    for day in active_days:
        sale_date = date(YEAR, month, day).isoformat()
        print(f"\n  Bills: {sale_date}")
        r = subprocess.run(
            ['python3', GENERATE_SCRIPT, '--date', sale_date],
            capture_output=True, text=True
        )
        print(r.stdout.rstrip())
        if r.returncode != 0:
            print(f"  ERROR: {r.stderr.rstrip()}")


if __name__ == '__main__':
    main()
