"""
Seed script — April 1, 2026 data only.
Inserts: ProductCategory, Product, Shop, Counter, CounterStaffAssignment,
         CounterSellingPrice, GodownStock, DailyCounterStock (C1/C2/C3),
         DailyCounterSummary, DailyExpense, CashRecord
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from datetime import date

conn = psycopg2.connect(
    host="195.201.119.186", port=5432,
    dbname="imfl_app", user="postgres", password="P@ssword4postgres"
)
conn.autocommit = False
cur = conn.cursor()

XLSX = "/home/prasadkabadi/ALK/Permits/KLS-I-APR-2026.xlsx"
APR1 = date(2026, 4, 1)

# ── Helpers ───────────────────────────────────────────────────

def safe_int(v, default=0):
    try:
        return int(float(v)) if v is not None else default
    except:
        return default

def safe_dec(v, default="0"):
    try:
        return str(round(float(v), 2)) if v is not None else default
    except:
        return default

def infer_category(short_name, master_code):
    n = short_name.lower()
    if "brandy" in n:
        return ("IMFL", "BRANDY")
    if any(x in n for x in ["whishy", "whisky", "whiskey"]):
        return ("IMFL", "WHISKY")
    if "rum" in n:
        return ("IMFL", "RUM")
    if "gin" in n:
        return ("IMFL", "GIN")
    if "vodka" in n:
        return ("IMFL", "VODKA")
    if "wine" in n:
        return ("IMFL", "WINE")
    if master_code >= 361:
        return ("BEER", None)
    return ("IMFL", "OTHERS")


# ═══════════════════════════════════════════════════════════════
# STEP 1 — ProductCategory
# ═══════════════════════════════════════════════════════════════
print("▶ Inserting ProductCategories...")

categories = [
    ("IMFL", "BRANDY",  None,      "Brandy"),
    ("IMFL", "WHISKY",  None,      "Whisky"),
    ("IMFL", "RUM",     None,      "Rum"),
    ("IMFL", "GIN",     None,      "Gin"),
    ("IMFL", "VODKA",   None,      "Vodka"),
    ("IMFL", "WINE",    None,      "Wine"),
    ("IMFL", "OTHERS",  None,      "Others (IMFL)"),
    ("BEER", None,      "BOTTLE",  "Beer Bottle"),
    ("BEER", None,      "TIN",     "Beer Tin"),
]

for product_type, imfl_cat, beer_pkg, display in categories:
    cur.execute("""
        INSERT INTO "ProductCategory" (product_type, imfl_category, beer_packaging, display_name)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT DO NOTHING
    """, (product_type, imfl_cat, beer_pkg, display))

cat_id = {}
cur.execute('SELECT id, product_type, imfl_category, beer_packaging FROM "ProductCategory"')
for row in cur.fetchall():
    cat_id[(row[1], row[2], row[3])] = row[0]

print(f"   Categories: {len(cat_id)}")


# ═══════════════════════════════════════════════════════════════
# STEP 2 — Products from Rate Matrix
# ═══════════════════════════════════════════════════════════════
print("▶ Reading Rate Matrix...")

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws_rate = wb["Rate Matrix"]

products = []
sp_map   = {}  # master_code -> selling_price on April 1

for i, row in enumerate(ws_rate.iter_rows(values_only=True)):
    if i < 2:
        continue
    master_code = row[0]
    brand       = row[1]
    ml          = row[2]
    ksbcl_name  = row[3]
    ksbcl_code  = row[4]
    carton_size = row[5]
    aroed       = row[6]
    sp_apr1     = row[10]  # S1 column

    if not master_code or not brand or not ml or not carton_size or carton_size == 0:
        continue

    is_tin = False
    if isinstance(ml, str) and ml.startswith("T"):
        ml_val = safe_int(ml[1:])
        is_tin = True
    else:
        try:
            ml_val = int(float(str(ml)))
        except:
            continue

    try:
        mc        = int(float(str(master_code)))
        cs        = int(float(str(carton_size)))
        aroed_val = round(float(aroed), 2) if aroed else 0.0
    except:
        continue

    ptype, imfl_cat = infer_category(brand.strip(), mc)
    cat_key = ("BEER", None, "TIN" if is_tin else "BOTTLE") if ptype == "BEER" else ("IMFL", imfl_cat, None)
    category_id = cat_id.get(cat_key)
    if not category_id:
        continue

    products.append({
        "master_code":      mc,
        "short_name":       brand.strip(),
        "ksbcl_item_name":  ksbcl_name.strip() if ksbcl_name else None,
        "ksbcl_item_code":  str(int(float(str(ksbcl_code)))) if ksbcl_code else None,
        "category_id":      category_id,
        "volume_ml":        ml_val,
        "carton_size":      cs,
        "aroed_per_bottle": aroed_val,
    })

    if sp_apr1 is not None:
        sp_map[mc] = float(sp_apr1)

print(f"▶ Inserting {len(products)} Products...")

for p in products:
    cur.execute("""
        INSERT INTO "Product"
            (master_code, short_name, ksbcl_item_name, ksbcl_item_code,
             category_id, volume_ml, carton_size, aroed_per_bottle,
             is_active, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, true, now(), now())
        ON CONFLICT (master_code) DO NOTHING
    """, (p["master_code"], p["short_name"], p["ksbcl_item_name"], p["ksbcl_item_code"],
          p["category_id"], p["volume_ml"], p["carton_size"], p["aroed_per_bottle"]))

prod_id = {}
cur.execute('SELECT id, master_code FROM "Product"')
for row in cur.fetchall():
    prod_id[row[1]] = row[0]

print(f"   Products in DB: {len(prod_id)}")


# ═══════════════════════════════════════════════════════════════
# STEP 3 — Shop & Counters
# ═══════════════════════════════════════════════════════════════
print("▶ Inserting Shop KLS...")

cur.execute("""
    INSERT INTO "Shop" (name, short_name, address, ksbcl_retailer_id,
                        owner_name, is_active, created_at, updated_at)
    VALUES ('Kabadi Liquor Shop', 'KLS', 'H.No.406/2, B3, Bazar Galli, Vadgaon',
            '16884', 'Prasad Ashok Kabadi', true, now(), now())
    ON CONFLICT (short_name) DO NOTHING
""")
cur.execute('SELECT id FROM "Shop" WHERE short_name = %s', ('KLS',))
shop_id = cur.fetchone()[0]
print(f"   Shop ID: {shop_id}")

print("▶ Inserting Counters...")
for order, name in [(1, "Counter 1"), (2, "Counter 2"), (3, "Counter 3")]:
    cur.execute("""
        INSERT INTO "Counter" (shop_id, name, display_order, is_active)
        VALUES (%s, %s, %s, true)
        ON CONFLICT (shop_id, display_order) DO NOTHING
    """, (shop_id, name, order))

counter_id = {}
cur.execute('SELECT id, display_order FROM "Counter" WHERE shop_id = %s ORDER BY display_order', (shop_id,))
for row in cur.fetchall():
    counter_id[row[1]] = row[0]
print(f"   Counters: {counter_id}")


# ═══════════════════════════════════════════════════════════════
# STEP 4 — CounterStaffAssignment
# ═══════════════════════════════════════════════════════════════
print("▶ Inserting Staff Assignments...")

for cnum, name in [(1, "Ramesh"), (3, "Sudesh")]:
    cur.execute("""
        INSERT INTO "CounterStaffAssignment" (counter_id, staff_name, assigned_date)
        VALUES (%s, %s, %s)
    """, (counter_id[cnum], name, APR1))


# ═══════════════════════════════════════════════════════════════
# STEP 5 — CounterSellingPrice (S1 from Rate Matrix)
# ═══════════════════════════════════════════════════════════════
print("▶ Inserting CounterSellingPrices...")

sp_rows = []
for mc, sp in sp_map.items():
    pid = prod_id.get(mc)
    if not pid:
        continue
    for cnum in [1, 2, 3]:
        sp_rows.append((counter_id[cnum], pid, str(sp), APR1))

execute_values(cur, """
    INSERT INTO "CounterSellingPrice" (counter_id, product_id, selling_price, effective_from)
    SELECT v.counter_id, v.product_id, v.selling_price::numeric, v.effective_from::date
    FROM (VALUES %s) AS v(counter_id, product_id, selling_price, effective_from)
""", sp_rows)
print(f"   SP records: {len(sp_rows)}")


# ═══════════════════════════════════════════════════════════════
# STEP 6 — Read Sheet "1" and insert stock data
# ═══════════════════════════════════════════════════════════════
print("▶ Reading Sheet '1' (April 1)...")

ws1 = wb["1"]

def get_counter_cols(row, base):
    return {
        "ob":     safe_int(row[base + 2]),
        "rec":    safe_int(row[base + 3]),
        "total":  safe_int(row[base + 4]),
        "sale":   safe_int(row[base + 5]),
        "cb":     safe_int(row[base + 6]),
        "amt":    safe_dec(row[base + 7]),
        "tips":   safe_dec(row[base + 8]),
        "totamt": safe_dec(row[base + 9]),
        "mrp":    safe_dec(row[base + 10]),
        "sp":     safe_dec(row[base + 11]),
    }

def get_godown_cols(row):
    return {
        "ob":  safe_int(row[59]),
        "rec": safe_int(row[60]),
        "sc1": safe_int(row[62]),
        "sc2": safe_int(row[63]),
        "sc3": safe_int(row[64]),
        "cb":  safe_int(row[65]),
    }

stock_c = {1: [], 2: [], 3: []}
stock_gd = []

for i, row in enumerate(ws1.iter_rows(values_only=True)):
    if i < 2 or i >= 420:
        continue
    mc = row[0]
    if mc is None or not isinstance(mc, (int, float)):
        continue
    mc  = int(mc)
    pid = prod_id.get(mc)
    if not pid:
        continue

    stock_c[1].append((pid, get_counter_cols(row, 1)))
    stock_c[2].append((pid, get_counter_cols(row, 15)))
    stock_c[3].append((pid, get_counter_cols(row, 29)))
    stock_gd.append((pid, get_godown_cols(row)))

print(f"   Product rows found: {len(stock_c[1])}")

# ── 6a: DailyCounterStock ─────────────────────────────────────
print("▶ Inserting DailyCounterStock...")

for cnum in [1, 2, 3]:
    rows = []
    for pid, d in stock_c[cnum]:
        rows.append((counter_id[cnum], pid, APR1,
                     d["ob"], d["rec"], d["total"], d["sale"], d["cb"],
                     d["mrp"], d["sp"], d["amt"], d["tips"], d["totamt"]))
    execute_values(cur, """
        INSERT INTO "DailyCounterStock"
            (counter_id, product_id, stock_date,
             opening_balance_btls, received_from_godown_btls, total_btls,
             sold_btls, closing_balance_btls,
             mrp_per_bottle, selling_price_per_bottle,
             sale_amount_mrp, tips_amount, total_sale_amount)
        SELECT v.cid, v.pid, v.dt::date,
               v.ob::int, v.rec::int, v.total::int,
               v.sale::int, v.cb::int,
               v.mrp::numeric, v.sp::numeric,
               v.amt::numeric, v.tips::numeric, v.totamt::numeric
        FROM (VALUES %s) AS v(cid, pid, dt, ob, rec, total, sale, cb, mrp, sp, amt, tips, totamt)
    """, rows)

print(f"   DailyCounterStock rows: {len(stock_c[1]) * 3}")

# ── 6b: GodownStock ───────────────────────────────────────────
print("▶ Inserting GodownStock...")

gd_rows = [(shop_id, pid, APR1, gd["ob"], gd["rec"], gd["cb"]) for pid, gd in stock_gd]
execute_values(cur, """
    INSERT INTO "GodownStock" (shop_id, product_id, stock_date,
                               opening_balance_btls, received_btls, closing_balance_btls)
    SELECT v.shop_id, v.pid, v.dt::date, v.ob::int, v.rec::int, v.cb::int
    FROM (VALUES %s) AS v(shop_id, pid, dt, ob, rec, cb)
    ON CONFLICT (shop_id, product_id, stock_date) DO NOTHING
""", gd_rows)

cur.execute('SELECT id, product_id FROM "GodownStock" WHERE shop_id = %s AND stock_date = %s',
            (shop_id, APR1))
gd_stock_id = {row[1]: row[0] for row in cur.fetchall()}

# ── 6c: GodownDistribution ────────────────────────────────────
print("▶ Inserting GodownDistributions...")

dist_rows = []
for pid, gd in stock_gd:
    gsid = gd_stock_id.get(pid)
    if not gsid:
        continue
    for cnum, qty in [(1, gd["sc1"]), (2, gd["sc2"]), (3, gd["sc3"])]:
        if qty > 0:
            dist_rows.append((gsid, counter_id[cnum], qty, APR1))

if dist_rows:
    execute_values(cur, """
        INSERT INTO "GodownDistribution" (godown_stock_id, counter_id, distributed_btls, distribution_date)
        SELECT v.gsid, v.cid, v.qty::int, v.dt::date
        FROM (VALUES %s) AS v(gsid, cid, qty, dt)
    """, dist_rows)

print(f"   GodownDistribution rows: {len(dist_rows)}")


# ═══════════════════════════════════════════════════════════════
# STEP 7 — DailyCounterSummary
# ═══════════════════════════════════════════════════════════════
print("▶ Inserting DailyCounterSummary...")

summaries = [
    (1, "Ramesh", "24325", "1860", "505", "26185", "2520", "2610", "17650", "5420", "505",  "0"),
    (2, None,     "39945", "1780", "681", "41725", "6505", "6595", "33000", "1449", "681",  "0"),
    (3, "Sudesh", "30265", "2085", "577", "32350", "3290", "3660", "24200", "3910", "577", "-3"),
]

for cnum, staff, liq, beer, tips, grand, gpay, exp, coll, drawer, tips_cash, tally in summaries:
    cur.execute("""
        INSERT INTO "DailyCounterSummary"
            (counter_id, summary_date, staff_name,
             liquor_sale_amount, beer_sale_amount, total_tips, grand_total_by_sale,
             google_pay_amount, expenses_total, collection_total,
             drawer_cash_total, tips_cash_total, tally_difference)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (counter_id, summary_date) DO NOTHING
    """, (counter_id[cnum], APR1, staff,
          liq, beer, tips, grand, gpay, exp, coll, drawer, tips_cash, tally))


# ═══════════════════════════════════════════════════════════════
# STEP 8 — DailyExpense
# ═══════════════════════════════════════════════════════════════
print("▶ Inserting DailyExpenses...")

expenses_data = [
    (1, "TEA",        40,   False),
    (1, "GOOGLE_PAY", 2520, True),
    (1, "BHATTA",     50,   False),
    (2, "TEA",        40,   False),
    (2, "GOOGLE_PAY", 6505, True),
    (2, "BHATTA",     50,   False),
    (3, "TEA",        40,   False),
    (3, "BREAKAGE",   280,  False),
    (3, "GOOGLE_PAY", 3290, True),
    (3, "BHATTA",     50,   False),
]

for cnum, category, amount, is_upi in expenses_data:
    cur.execute("""
        INSERT INTO "DailyExpense" (counter_id, expense_date, category, amount, is_upi)
        VALUES (%s, %s, %s, %s, %s)
    """, (counter_id[cnum], APR1, category, str(amount), is_upi))


# ═══════════════════════════════════════════════════════════════
# STEP 9 — CashRecord
# ═══════════════════════════════════════════════════════════════
print("▶ Inserting CashRecords...")

cash_records = [
    (1, "COLLECTION", "NOTE", 500,  18),
    (1, "COLLECTION", "NOTE", 200,   5),
    (1, "COLLECTION", "NOTE", 100,  45),
    (1, "COLLECTION", "NOTE",  50,  50),
    (1, "COLLECTION", "NOTE",  20,  25),
    (1, "COLLECTION", "NOTE",  10,  15),
    (1, "COLLECTION", "COIN",   1, 505),
    (2, "COLLECTION", "NOTE", 500,  42),
    (2, "COLLECTION", "NOTE", 200,  37),
    (2, "COLLECTION", "NOTE", 100,  32),
    (2, "COLLECTION", "NOTE",  50,  28),
    (2, "COLLECTION", "COIN",   1, 681),
    (3, "COLLECTION", "NOTE", 500,  21),
    (3, "COLLECTION", "NOTE", 200,  21),
    (3, "COLLECTION", "NOTE", 100,  70),
    (3, "COLLECTION", "NOTE",  50,  50),
    (3, "COLLECTION", "COIN",   1, 577),
    (1, "DRAWER_CASH", "COIN", 1, 5420),
    (2, "DRAWER_CASH", "COIN", 1, 1449),
    (3, "DRAWER_CASH", "COIN", 1, 3910),
    (1, "TIPS_CASH", "COIN", 1, 505),
    (2, "TIPS_CASH", "COIN", 1, 681),
    (3, "TIPS_CASH", "COIN", 1, 577),
]

for cnum, cat, dtype, dval, count in cash_records:
    cur.execute("""
        INSERT INTO "CashRecord"
            (counter_id, record_date, cash_category, denomination_type,
             denomination_value, count, total_amount)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (counter_id, record_date, cash_category, denomination_type, denomination_value)
        DO NOTHING
    """, (counter_id[cnum], APR1, cat, dtype, dval, count, dval * count))


# ═══════════════════════════════════════════════════════════════
# COMMIT
# ═══════════════════════════════════════════════════════════════
conn.commit()
cur.close()
conn.close()

print()
print("✓ April 1, 2026 seed complete.")
print(f"  Products          : {len(prod_id)}")
print(f"  DailyCounterStock : {len(stock_c[1]) * 3} rows")
print(f"  GodownStock       : {len(gd_rows)} rows")
print(f"  GodownDistributions: {len(dist_rows)} rows")
print(f"  DailyExpenses     : {len(expenses_data)} rows")
print(f"  CashRecords       : {len(cash_records)} rows")
