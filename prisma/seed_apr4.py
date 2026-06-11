"""
Seed script — April 4, 2026 (Day 4).
Inserts: GodownStock, DailyCounterStock (C1/C2/C3),
         StockLot (source=PERMIT, 35 products, 6408 btls),
         DailyCounterSummary, DailyExpense, CashRecord

Column layout note:
  - Godown section: ob=col59, rec=col60, sc1=62, sc2=63, sc3=64, cb=col65
  - Shop section:   ob=col45, rec=col46(AU)=PERMIT RECEIVED, total=47, sale=48, cb=49
  On permit days, col60 (godown Rec) is 0 — the permit receipt is tracked at
  shop level in col46 (AU). Use col46 for GodownStock.received_btls and
  StockLot.initial_qty. MRP from Rate Matrix M4 (col15).

  C3 staff: Jadhav.  No MRP/SP changes from Day 3.
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from decimal import Decimal
from datetime import date
from urllib.parse import urlparse, unquote
import os

raw_url = os.environ.get('DATABASE_URL',
    'postgresql://postgres:P%40ssword4postgres@195.201.119.186:5432/imfl_app')
raw_url = raw_url.split('?')[0]
u = urlparse(raw_url)
conn = psycopg2.connect(
    host=u.hostname, port=u.port or 5432,
    dbname=u.path.lstrip('/'), user=u.username, password=unquote(u.password),
)
conn.autocommit = False
cur = conn.cursor()

XLSX = '/home/prasadkabadi/ALK/Permits/KLS-I-APR-2026.xlsx'
APR4 = date(2026, 4, 4)

def safe_int(v, default=0):
    try: return int(float(v)) if v is not None else default
    except: return default

def safe_dec(v, default='0'):
    try: return str(round(float(v), 2)) if v is not None else default
    except: return default

# ── References ─────────────────────────────────────────────────
cur.execute('SELECT id FROM "Shop" WHERE short_name = %s', ('KLS',))
shop_id = cur.fetchone()[0]
cur.execute('SELECT id, master_code FROM "Product"')
prod_id = {mc: pid for pid, mc in cur.fetchall()}
cur.execute('SELECT id, display_order FROM "Counter" WHERE shop_id = %s ORDER BY display_order', (shop_id,))
counter_id = {order: cid for cid, order in cur.fetchall()}
print(f'Shop={shop_id}, Products={len(prod_id)}, Counters={counter_id}')

# ── Rate Matrix M4 MRP (col15) ─────────────────────────────────
wb     = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws_rm  = wb['Rate Matrix']
rate_m4 = {}
for row in ws_rm.iter_rows(min_row=3, values_only=True):
    mc = row[0]
    if mc is None: continue
    mrp = row[15]  # M4
    if mrp is not None:
        rate_m4[int(float(mc))] = Decimal(str(mrp))

# ── Read sheet "4" ─────────────────────────────────────────────
ws4 = wb['4']

def counter_cols(row, base):
    return {
        'ob':     safe_int(row[base + 2]),
        'rec':    safe_int(row[base + 3]),
        'total':  safe_int(row[base + 4]),
        'sale':   safe_int(row[base + 5]),
        'cb':     safe_int(row[base + 6]),
        'amt':    safe_dec(row[base + 7]),
        'tips':   safe_dec(row[base + 8]),
        'totamt': safe_dec(row[base + 9]),
        'mrp':    safe_dec(row[base + 10]),
        'sp':     safe_dec(row[base + 11]),
    }

def godown_cols(row):
    return {
        'ob':  safe_int(row[59]),
        'rec': safe_int(row[46]),   # col46 (AU) = shop-level permit received
        'sc1': safe_int(row[62]),
        'sc2': safe_int(row[63]),
        'sc3': safe_int(row[64]),
        'cb':  safe_int(row[65]),
    }

stock_c  = {1: [], 2: [], 3: []}
stock_gd = []
permit_lots = []   # (mc, qty, mrp) for StockLot creation

for i, row in enumerate(ws4.iter_rows(values_only=True)):
    if i < 2 or i >= 420:
        continue
    mc = row[0]
    if mc is None or not isinstance(mc, (int, float)):
        continue
    mc  = int(mc)
    pid = prod_id.get(mc)
    if not pid:
        continue

    stock_c[1].append((pid, counter_cols(row, 1)))
    stock_c[2].append((pid, counter_cols(row, 15)))
    stock_c[3].append((pid, counter_cols(row, 29)))
    stock_gd.append((pid, godown_cols(row)))

    permit_qty = safe_int(row[46])
    if permit_qty > 0:
        mrp = rate_m4.get(mc)
        if mrp:
            permit_lots.append((pid, permit_qty, mrp))

print(f'Product rows in sheet: {len(stock_c[1])}')
print(f'Permit lots to create: {len(permit_lots)} products, {sum(q for _,q,_ in permit_lots)} btls')

# ── DailyCounterStock ──────────────────────────────────────────
print('▶ Inserting DailyCounterStock...')
for cnum in [1, 2, 3]:
    rows = []
    for pid, d in stock_c[cnum]:
        rows.append((counter_id[cnum], pid, APR4,
                     d['ob'], d['rec'], d['total'], d['sale'], d['cb'],
                     d['mrp'], d['sp'], d['amt'], d['tips'], d['totamt']))
    execute_values(cur, """
        INSERT INTO "DailyCounterStock"
            (counter_id, product_id, stock_date,
             opening_balance_btls, received_from_godown_btls, total_btls,
             sold_btls, closing_balance_btls,
             mrp_per_bottle, selling_price_per_bottle,
             sale_amount_mrp, tips_amount, total_sale_amount)
        SELECT v.cid, v.pid, v.dt::date,
               v.ob::int, v.rec::int, v.total::int,
               v.sale::int, v.cb::int,
               v.mrp::numeric, v.sp::numeric,
               v.amt::numeric, v.tips::numeric, v.totamt::numeric
        FROM (VALUES %s) AS v(cid,pid,dt,ob,rec,total,sale,cb,mrp,sp,amt,tips,totamt)
        ON CONFLICT DO NOTHING
    """, rows)
print(f'   DailyCounterStock rows: {len(stock_c[1]) * 3}')

# ── GodownStock ────────────────────────────────────────────────
print('▶ Inserting GodownStock...')
gd_rows = [(shop_id, pid, APR4, gd['ob'], gd['rec'], gd['cb']) for pid, gd in stock_gd]
execute_values(cur, """
    INSERT INTO "GodownStock" (shop_id, product_id, stock_date,
                               opening_balance_btls, received_btls, closing_balance_btls)
    SELECT v.shop_id, v.pid, v.dt::date, v.ob::int, v.rec::int, v.cb::int
    FROM (VALUES %s) AS v(shop_id, pid, dt, ob, rec, cb)
    ON CONFLICT (shop_id, product_id, stock_date) DO NOTHING
""", gd_rows)

cur.execute('SELECT id, product_id FROM "GodownStock" WHERE shop_id = %s AND stock_date = %s',
            (shop_id, APR4))
gd_stock_id = {row[1]: row[0] for row in cur.fetchall()}

# ── GodownDistribution (from godown SC columns — 0 on permit day) ──
print('▶ Inserting GodownDistributions...')
dist_rows = []
for pid, gd in stock_gd:
    gsid = gd_stock_id.get(pid)
    if not gsid:
        continue
    for cnum, qty in [(1, gd['sc1']), (2, gd['sc2']), (3, gd['sc3'])]:
        if qty > 0:
            dist_rows.append((gsid, counter_id[cnum], qty, APR4))
if dist_rows:
    execute_values(cur, """
        INSERT INTO "GodownDistribution" (godown_stock_id, counter_id, distributed_btls, distribution_date)
        SELECT v.gsid, v.cid, v.qty::int, v.dt::date
        FROM (VALUES %s) AS v(gsid, cid, qty, dt)
    """, dist_rows)
print(f'   GodownDistribution rows: {len(dist_rows)}')

# ── StockLot (PERMIT) ──────────────────────────────────────────
print('▶ Inserting StockLots (PERMIT)...')

# Idempotency: skip products that already have a PERMIT lot on this date
cur.execute("""
    SELECT product_id FROM "StockLot"
    WHERE shop_id = %s AND lot_date = %s AND source = 'PERMIT'
""", (shop_id, APR4))
already_done = {r[0] for r in cur.fetchall()}

lot_rows = [
    (shop_id, pid, mrp, qty, 0, APR4)
    for pid, qty, mrp in permit_lots
    if pid not in already_done
]

if lot_rows:
    cur.executemany("""
        INSERT INTO "StockLot"
            (shop_id, product_id, mrp_per_bottle, initial_qty, consumed_qty,
             lot_date, source, permit_item_id)
        VALUES (%s, %s, %s, %s, %s, %s, 'PERMIT', NULL)
    """, lot_rows)
print(f'   StockLot rows: {len(lot_rows)}')

# ── DailyCounterSummary ────────────────────────────────────────
print('▶ Inserting DailyCounterSummary...')
summaries = [
    (1, 'Ramesh', '21379', '2395', '401', '23774', '3935',  '5695', '15850', '2630', '401',  '1'),
    (2, None,     '30421',  '645', '469', '31066', '6805',  '6895', '23489', '1151', '489',  '0'),
    (3, 'Jadhav', '36115',  '565', '810', '36680', '7070',  '7160', '27810', '2520', '810',  '0'),
]
for cnum, staff, liq, beer, tips, grand, gpay, exp, coll, drawer, tips_cash, tally in summaries:
    cur.execute("""
        INSERT INTO "DailyCounterSummary"
            (counter_id, summary_date, staff_name,
             liquor_sale_amount, beer_sale_amount, total_tips, grand_total_by_sale,
             google_pay_amount, expenses_total, collection_total,
             drawer_cash_total, tips_cash_total, tally_difference)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (counter_id, summary_date) DO NOTHING
    """, (counter_id[cnum], APR4, staff,
          liq, beer, tips, grand, gpay, exp, coll, drawer, tips_cash, tally))

# ── DailyExpense ───────────────────────────────────────────────
print('▶ Inserting DailyExpenses...')
expenses_data = [
    (1, 'TEA',         40,    False),
    (1, 'PERMIT_RENT', 1400,  False),
    (1, 'BREAKAGE',    250,   False),
    (1, 'GOOGLE_PAY',  3935,  True),
    (1, 'BHATTA',      50,    False),
    (1, 'PAK',         20,    False),
    (2, 'TEA',         40,    False),
    (2, 'GOOGLE_PAY',  6805,  True),
    (2, 'BHATTA',      50,    False),
    (3, 'TEA',         40,    False),
    (3, 'GOOGLE_PAY',  7070,  True),
    (3, 'BHATTA',      50,    False),
]
for cnum, category, amount, is_upi in expenses_data:
    cur.execute("""
        INSERT INTO "DailyExpense" (counter_id, expense_date, category, amount, is_upi)
        VALUES (%s, %s, %s, %s, %s)
    """, (counter_id[cnum], APR4, category, str(amount), is_upi))

# ── CashRecord ─────────────────────────────────────────────────
print('▶ Inserting CashRecords...')
cash_records = [
    (1, 'COLLECTION', 'NOTE', 500,  11),
    (1, 'COLLECTION', 'NOTE', 200,   5),
    (1, 'COLLECTION', 'NOTE', 100,  55),
    (1, 'COLLECTION', 'NOTE',  50,  50),
    (1, 'COLLECTION', 'NOTE',  20,  40),
    (1, 'COLLECTION', 'NOTE',  10,  15),
    (2, 'COLLECTION', 'NOTE', 500,  32),
    (2, 'COLLECTION', 'NOTE', 200,  13),
    (2, 'COLLECTION', 'NOTE', 100,  37),
    (2, 'COLLECTION', 'NOTE',  50,  14),
    (3, 'COLLECTION', 'NOTE', 500,  45),
    (3, 'COLLECTION', 'NOTE', 200,   5),
    (3, 'COLLECTION', 'NOTE', 100,  20),
    (3, 'COLLECTION', 'NOTE',  50,  30),
    (1, 'DRAWER_CASH', 'COIN', 1, 2630),
    (2, 'DRAWER_CASH', 'COIN', 1, 1151),
    (3, 'DRAWER_CASH', 'COIN', 1, 2520),
    (1, 'TIPS_CASH', 'COIN', 1, 401),
    (2, 'TIPS_CASH', 'COIN', 1, 489),
    (3, 'TIPS_CASH', 'COIN', 1, 810),
]
for cnum, cat, dtype, dval, count in cash_records:
    cur.execute("""
        INSERT INTO "CashRecord"
            (counter_id, record_date, cash_category, denomination_type,
             denomination_value, count, total_amount)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (counter_id, record_date, cash_category, denomination_type, denomination_value)
        DO NOTHING
    """, (counter_id[cnum], APR4, cat, dtype, dval, count, dval * count))

# ── Commit ─────────────────────────────────────────────────────
conn.commit()
cur.close()
conn.close()

print()
print('✓ April 4, 2026 seed complete.')
print(f'  DailyCounterStock   : {len(stock_c[1]) * 3} rows')
print(f'  GodownStock         : {len(gd_rows)} rows (permit received_btls from col46/AU)')
print(f'  GodownDistributions : {len(dist_rows)} rows')
print(f'  StockLot (PERMIT)   : {len(lot_rows)} rows, {sum(q for _,q,_ in permit_lots)} btls, MRP from M4')
print(f'  DailyExpenses       : {len(expenses_data)} rows')
print(f'  CashRecords         : {len(cash_records)} rows')
