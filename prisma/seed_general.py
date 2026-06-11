#!/usr/bin/env python3
"""
seed_general.py — Generalised daily seeder for KLS May-Jun 2026.

Usage:
  python3 seed_general.py --date 2026-05-14
  python3 seed_general.py --start 2026-05-14 --end 2026-05-30
  python3 seed_general.py --start 2026-05-14 --end 2026-05-30 --dry-run
  python3 seed_general.py --start 2026-05-14 --end 2026-05-30 --no-bills

Received stock for the godown is taken from the permit PDFs (CNF CBS × carton),
not from Excel column AU (shop consolidated Rec).
"""

import argparse, os, subprocess, sys
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── DB ─────────────────────────────────────────────────────────────────────────
CONN = dict(host="195.201.119.186", port=5432,
            dbname="imfl_app", user="postgres", password="P@ssword4postgres")
SHOP  = "KLS"
YEAR  = 2026
GENERATE_SCRIPT = os.path.join(os.path.dirname(__file__), "generate_daily_bills.py")

# ── Excel files ────────────────────────────────────────────────────────────────
XLSX = {
    "I":  "/home/prasadkabadi/ALK/Permits/KLS-I-MAY-2026.xlsx",    # sheets 1-15
    "II": "/home/prasadkabadi/ALK/Permits/KLS-II-MAY-2026.xlsx",   # sheets 16-30
}

def xlsx_for_day(day: int) -> str:
    return XLSX["I"] if day <= 15 else XLSX["II"]

# ── Permit schedule: {date_str: {master_code: bottles}} ───────────────────────
# CNF CBS × carton_size from KSBCL Indent Report PDFs.
# Only dates that had an actual delivery are listed.
# Days with no entry → zero received (stock stays in godown from prev permit).

PERMIT_REC = {
    # 13-May already seeded via seed_may13.py — not repeated here.

    # ── INDBG226002026  14-May ─────────────────────────────────────────────────
    "2026-05-14": {
        185: 12*48,   # Green Champion 180ML
        201:  2*96,   # Green Champion 90ML
        366: 10*12,   # KF Strong 650ML
        369:  1*24,   # KF Strong CAN 500ML
        404: 20*12,   # Bullet Strong 650ML
        383:  2*12,   # Budweiser Magnum 650ML
        384:  1*24,   # Budweiser Magnum CAN 500ML
        390:  2*24,   # RC Strong CAN 500ML
        388: 15*12,   # RC Strong 650ML
        376: 10*12,   # Tuborg Strong 650ML
        406:  1*24,   # Bullet Strong 330ML
        405:  3*24,   # Bullet Strong CAN 500ML
        386:  3*12,   # Power Cool Strong 650ML
    },

    # ── INDBG226002184  18-May ─────────────────────────────────────────────────
    "2026-05-18": {
         30:  1*48,   # 8PM Special 180ML
         45:  3*48,   # Bagpiper Deluxe 180ML
         63:  2*48,   # DSP Black PET 180ML
         69:  1*48,   # Haywards Cheers 180ML
         83:  1*48,   # McDowell's No.1 Reserve 180ML
        108:  3*48,   # Old Tavern 180ML
        113:  1*48,   # OC Special 180ML
        114:  1*96,   # OC Special 90ML
        115:  1*12,   # ORC PET 750ML
        116:  1*24,   # ORC 375ML
        145: 15*48,   # Wellington 180ML
        146:  3*96,   # Wellington 90ML
        155:  3*48,   # DK 180ML
        156: 72*96,   # DK 90ML
        159:  3*96,   # MaQintosh 90ML
        162:  3*48,   # OC Star Supreme 180ML
        174:  3*48,   # Black Belt 180ML
        186:  1*48,   # Original Traveller 180ML
        187:  1*96,   # Original Traveller 90ML
        219:  1*48,   # Old Monk Rum 180ML
        234:  3*48,   # McDowell's Classic Rum 180ML
        255:  3*48,   # Carnival Dry Gin 180ML
        332:  1*48,   # Rico Red Wine 180ML
        363:  3*12,   # KF Premium Lager 650ML
        379:  1*12,   # Budweiser Premium 650ML
    },

    # ── INDBG226002372  21-May ─────────────────────────────────────────────────
    "2026-05-21": {
         31:  1*96,   # 8PM Special 90ML
         43:  1*12,   # Bagpiper Deluxe PET 750ML
         45:  1*48,   # Bagpiper Deluxe 180ML
         69:  1*48,   # Haywards Cheers 180ML
         73:  1*48,   # Imperial Blue 180ML
        109:  2*96,   # Old Tavern 90ML
        145:  6*48,   # Wellington 180ML
        146:  1*96,   # Wellington 90ML
        155:  6*48,   # DK 180ML
        174:  1*48,   # Black Belt 180ML
        185:  9*48,   # Green Champion 180ML
        201:  3*96,   # Green Champion 90ML
        206:  1*48,   # Bagpiper XXX Rum 180ML
        235:  1*96,   # McDowell's Classic Rum 90ML
        363: 10*12,   # KF Premium Lager 650ML
        366: 12*12,   # KF Strong 650ML
        367:  2*24,   # KF Strong 330ML
        383:  1*12,   # Budweiser Magnum 650ML
        386: 12*12,   # Power Cool Strong 650ML
    },

    # ── INDBG226002519  23-May ─────────────────────────────────────────────────
    "2026-05-23": {
         45:  2*48,   # Bagpiper Deluxe 180ML
         62:  1*24,   # DSP Black 375ML
         64:  1*96,   # DSP Black 90ML
         69:  1*48,   # Haywards Cheers 180ML
         72:  1*24,   # Imperial Blue 375ML
        115:  3*12,   # ORC PET 750ML
        117: 30*48,   # ORC ABP 180ML
        145:  6*48,   # Wellington 180ML
        151:  1*12,   # Black Belt PET 750ML
        155:  6*48,   # DK 180ML
        156:  6*96,   # DK 90ML
        174:  1*48,   # Black Belt 180ML
        187:  1*96,   # Original Traveller 90ML
        206:  1*48,   # Bagpiper XXX Rum 180ML
        219:  1*48,   # Old Monk Rum 180ML
        220:  1*96,   # Old Monk Rum 90ML
        221:  1*48,   # Raja Rum 180ML
        255:  2*48,   # Carnival Dry Gin 180ML
        256:  1*96,   # Carnival Dry Gin 90ML
        274:  1*48,   # Romanov Orange Vodka 180ML
        378:  3*24,   # Tuborg Strong CAN 500ML
        383:  1*12,   # Budweiser Magnum 650ML
    },

    # ── INDBG226002556  25-May  (RATIONED: 90→9 CBs) ──────────────────────────
    "2026-05-25": {
        404:  9*12,   # Bullet Strong 650ML  (ordered 90, confirmed 9)
    },

    # ── INDBG226002700  27-May  (RATIONED: Bagpiper 4→2 CBs) ──────────────────
    "2026-05-27": {
         30:  1*48,   # 8PM Special 180ML
         45:  2*48,   # Bagpiper Deluxe 180ML  (ordered 4, confirmed 2)
        108:  1*48,   # Old Tavern 180ML
        116:  2*24,   # ORC 375ML
        145:  9*48,   # Wellington 180ML
        146:  3*96,   # Wellington 90ML
        155:  9*48,   # DK 180ML
        174:  3*48,   # Black Belt 180ML
        187:  1*96,   # Original Traveller 90ML
        204:  1*96,   # Carnival Rum 90ML
        206:  1*48,   # Bagpiper XXX Rum 180ML
        255:  3*48,   # Carnival Dry Gin 180ML
        328:  1*48,   # Goan's Wine 180ML
        366: 15*12,   # KF Strong 650ML
        383:  1*12,   # Budweiser Magnum 650ML
        388: 15*12,   # RC Strong 650ML
        404: 15*12,   # Bullet Strong 650ML
    },

    # ── INDBG226002835  30-May ─────────────────────────────────────────────────
    "2026-05-30": {
         45:  3*48,   # Bagpiper Deluxe 180ML
         58:  1*48,   # Director's Special Tetra 180ML
         69:  1*48,   # Haywards Cheers 180ML
        108:  1*48,   # Old Tavern 180ML
        114:  1*96,   # OC Special 90ML
        155:  6*48,   # DK 180ML
        156: 39*96,   # DK 90ML
        174:  1*48,   # Black Belt 180ML
        203:  1*48,   # Carnival Rum 180ML
        219:  1*48,   # Old Monk Rum 180ML
        367:  1*24,   # KF Strong 330ML
        369:  3*24,   # KF Strong CAN 500ML
        376: 15*12,   # Tuborg Strong 650ML
        379:  1*12,   # Budweiser Premium 650ML
        383:  1*12,   # Budweiser Magnum 650ML
        386:  6*12,   # Power Cool Strong 650ML
        390:  2*24,   # RC Strong CAN 500ML
        404:  9*12,   # Bullet Strong 650ML
    },
}

# ── Column layout (0-indexed) ──────────────────────────────────────────────────
CTR_BASES = {1: 1, 2: 15, 3: 29}
R_STAFF      = 423
R_LIQ        = 425
R_BEER       = 426
R_GTOTAL_ROW = 444
R_GPAY       = 439
R_EXP_TOTAL  = 445
R_TIPS_CASH  = 460
EXPENSE_ROWS = [
    (433, 'TEA', False), (434, 'PERMIT_RENT', False), (435, 'POOJA', False),
    (436, 'BREAKAGE', False), (437, 'OVER_CASH', False),
    (438, 'ELECTRICITY_BILL', False), (439, 'GOOGLE_PAY', True),
    (440, 'BHATTA', False), (441, 'GOOGLE_PAY_NEGATIVE', False),
    (442, 'PAK', False), (444, 'OTHERS', False),
]
EXP_COLS       = {1: 2,  2: 16, 3: 30}
COLL_COLS      = {1: (5,6,7,8), 2: (19,20,21,22), 3: (33,34,35,36)}
DRAW_TOTAL_COL = {1: 13, 2: 27, 3: 41}
COLL_TOTAL_COL = {1:  8, 2: 22, 3: 36}
TIPS_TOTAL_COL = {1:  8, 2: 22, 3: 36}
LIQ_SALE_COL   = {1:  6, 2: 20, 3: 34}
LIQ_TIPS_COL   = {1:  7, 2: 21, 3: 35}
BEER_SALE_COL  = {1:  6, 2: 20, 3: 34}
BEER_TIPS_COL  = {1:  7, 2: 21, 3: 35}
TALLY_COL      = {1: 11, 2: 25, 3: 39}
GPAY_COL       = {1:  2, 2: 16, 3: 30}
STAFF_COL      = {1:  1, 2: 15, 3: 29}


def si(v, d=0):
    try: return int(float(v)) if v is not None else d
    except: return d

def sd(v, d='0'):
    try: return str(round(float(v), 2)) if v is not None else d
    except: return d

def sfloat(v, d=0.0):
    try: return float(v) if v is not None else d
    except: return d


def extract_stock(rows, permit_rec):
    stock_c  = {1: [], 2: [], 3: []}
    stock_gd = []
    for i, row in enumerate(rows):
        if i < 2 or i >= 420:
            continue
        mc = row[0]
        if mc is None or not isinstance(mc, (int, float)):
            continue
        mc = int(mc)
        for cnum, base in CTR_BASES.items():
            stock_c[cnum].append((mc, {
                'ob':     si(row[base+2]), 'rec':    si(row[base+3]),
                'total':  si(row[base+4]), 'sale':   si(row[base+5]),
                'cb':     si(row[base+6]), 'amt':    sd(row[base+7]),
                'tips':   sd(row[base+8]), 'totamt': sd(row[base+9]),
                'mrp':    sd(row[base+10]), 'sp':    sd(row[base+11]),
            }))
        gd_rec = permit_rec.get(mc, si(row[60]))
        stock_gd.append((mc, {
            'ob':  si(row[59]), 'rec': gd_rec,
            'sc1': si(row[62]), 'sc2': si(row[63]),
            'sc3': si(row[64]), 'cb':  si(row[65]),
        }))
    return stock_c, stock_gd


def extract_summary(rows):
    result = {}
    for cnum in [1, 2, 3]:
        liq_tips  = sfloat(rows[R_LIQ][LIQ_TIPS_COL[cnum]])
        beer_tips = sfloat(rows[R_BEER][BEER_TIPS_COL[cnum]])
        result[cnum] = {
            'staff':      rows[R_STAFF][STAFF_COL[cnum]],
            'liq_sale':   sfloat(rows[R_LIQ][LIQ_SALE_COL[cnum]]),
            'beer_sale':  sfloat(rows[R_BEER][BEER_SALE_COL[cnum]]),
            'total_tips': liq_tips + beer_tips,
            'gpay':       sfloat(rows[R_GPAY][GPAY_COL[cnum]]),
            'exp_total':  sfloat(rows[R_EXP_TOTAL][EXP_COLS[cnum]]),
            'collection': sfloat(rows[R_GTOTAL_ROW][COLL_TOTAL_COL[cnum]]),
            'drawer':     sfloat(rows[R_GTOTAL_ROW][DRAW_TOTAL_COL[cnum]]),
            'tips_cash':  sfloat(rows[R_TIPS_CASH][TIPS_TOTAL_COL[cnum]]),
            'tally':      sfloat(rows[R_BEER][TALLY_COL[cnum]]),
        }
        s = result[cnum]
        s['grand'] = s['liq_sale'] + s['beer_sale']
    return result


def extract_expenses(rows):
    result = {1: [], 2: [], 3: []}
    for row_i, cat, is_upi in EXPENSE_ROWS:
        for cnum in [1, 2, 3]:
            amt = sfloat(rows[row_i][EXP_COLS[cnum]])
            if amt > 0:
                result[cnum].append((cat, amt, is_upi))
    return result


def extract_cash(rows):
    result = {1: [], 2: [], 3: []}
    for i in range(433, 444):
        row = rows[i]
        for cnum, (tc, dc, cc, totc) in COLL_COLS.items():
            dtype = row[tc]
            dval  = sfloat(row[dc])
            count = sfloat(row[cc])
            total = sfloat(row[totc])
            dtype_db = 'NOTE' if dtype=='Notes' else 'COIN' if dtype=='Coins' else None
            if dtype_db and count > 0 and total > 0:
                result[cnum].append(('COLLECTION', dtype_db, int(dval), int(count), int(total)))
    for cnum in [1, 2, 3]:
        d = int(sfloat(rows[R_GTOTAL_ROW][DRAW_TOTAL_COL[cnum]]))
        if d > 0:
            result[cnum].append(('DRAWER_CASH', 'COIN', 1, d, d))
    for cnum in [1, 2, 3]:
        t = int(sfloat(rows[R_TIPS_CASH][TIPS_TOTAL_COL[cnum]]))
        if t > 0:
            result[cnum].append(('TIPS_CASH', 'COIN', 1, t, t))
    return result


def seed_day(cur, shop_id, prod_id, counter_id, sale_date, rows, permit_rec, dry_run=False):
    stock_c, stock_gd = extract_stock(rows, permit_rec)
    summary            = extract_summary(rows)
    expenses           = extract_expenses(rows)
    cash               = extract_cash(rows)

    total_sold = sum(d['sale'] for cnum in [1,2,3] for _, d in stock_c[cnum])
    if total_sold == 0:
        print(f"  ⚠  No sales — skipping {sale_date}")
        return False

    rec_total = sum(permit_rec.values())
    print(f"\n  ── {sale_date}  sold:{total_sold}  permit_rec:{rec_total} btls ──")
    for cnum in [1, 2, 3]:
        s = summary[cnum]
        print(f"    C{cnum}: liq={s['liq_sale']:.0f}  beer={s['beer_sale']:.0f}  "
              f"tips={s['total_tips']:.0f}  gpay={s['gpay']:.0f}  tally={s['tally']:.0f}")
    if dry_run:
        return True

    # DailyCounterStock
    for cnum in [1, 2, 3]:
        rows_db = []
        for mc, d in stock_c[cnum]:
            pid = prod_id.get(mc)
            if not pid:
                continue
            rows_db.append((counter_id[cnum], pid, sale_date,
                            d['ob'], d['rec'], d['total'], d['sale'], d['cb'],
                            d['mrp'], d['sp'], d['amt'], d['tips'], d['totamt']))
        execute_values(cur, """
            INSERT INTO "DailyCounterStock"
                (counter_id, product_id, stock_date,
                 opening_balance_btls, received_from_godown_btls, total_btls,
                 sold_btls, closing_balance_btls,
                 mrp_per_bottle, selling_price_per_bottle,
                 sale_amount_mrp, tips_amount, total_sale_amount)
            SELECT v.cid, v.pid, v.dt::date,
                   v.ob::int, v.rec::int, v.total::int, v.sale::int, v.cb::int,
                   v.mrp::numeric, v.sp::numeric,
                   v.amt::numeric, v.tips::numeric, v.totamt::numeric
            FROM (VALUES %s) AS v(cid,pid,dt,ob,rec,total,sale,cb,mrp,sp,amt,tips,totamt)
            ON CONFLICT DO NOTHING
        """, rows_db)

    # GodownStock
    gd_rows = [(shop_id, prod_id[mc], sale_date, gd['ob'], gd['rec'], gd['cb'])
               for mc, gd in stock_gd if prod_id.get(mc)]
    execute_values(cur, """
        INSERT INTO "GodownStock" (shop_id, product_id, stock_date,
                                   opening_balance_btls, received_btls, closing_balance_btls)
        SELECT v.sid, v.pid, v.dt::date, v.ob::int, v.rec::int, v.cb::int
        FROM (VALUES %s) AS v(sid, pid, dt, ob, rec, cb)
        ON CONFLICT (shop_id, product_id, stock_date) DO NOTHING
    """, gd_rows)

    # GodownDistribution
    cur.execute('SELECT id, product_id FROM "GodownStock" WHERE shop_id=%s AND stock_date=%s',
                (shop_id, sale_date))
    gd_stock_id = {r[1]: r[0] for r in cur.fetchall()}
    dist_rows = []
    for mc, gd in stock_gd:
        pid  = prod_id.get(mc)
        gsid = gd_stock_id.get(pid)
        if not gsid:
            continue
        for cnum, qty in [(1, gd['sc1']), (2, gd['sc2']), (3, gd['sc3'])]:
            if qty > 0:
                dist_rows.append((gsid, counter_id[cnum], qty, sale_date))
    if dist_rows:
        execute_values(cur, """
            INSERT INTO "GodownDistribution"
                (godown_stock_id, counter_id, distributed_btls, distribution_date)
            SELECT v.gsid, v.cid, v.qty::int, v.dt::date
            FROM (VALUES %s) AS v(gsid, cid, qty, dt)
        """, dist_rows)

    # DailyCounterSummary
    for cnum in [1, 2, 3]:
        s = summary[cnum]
        cur.execute("""
            INSERT INTO "DailyCounterSummary"
                (counter_id, summary_date, staff_name,
                 liquor_sale_amount, beer_sale_amount, total_tips, grand_total_by_sale,
                 google_pay_amount, expenses_total, collection_total,
                 drawer_cash_total, tips_cash_total, tally_difference)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (counter_id, summary_date) DO NOTHING
        """, (counter_id[cnum], sale_date, s['staff'],
              str(s['liq_sale']), str(s['beer_sale']), str(s['total_tips']),
              str(s['grand']), str(s['gpay']), str(s['exp_total']),
              str(s['collection']), str(s['drawer']), str(s['tips_cash']),
              str(s['tally'])))

    # DailyExpense
    for cnum in [1, 2, 3]:
        for cat, amt, is_upi in expenses[cnum]:
            cur.execute("""
                INSERT INTO "DailyExpense" (counter_id, expense_date, category, amount, is_upi)
                VALUES (%s,%s,%s,%s,%s)
            """, (counter_id[cnum], sale_date, cat, str(amt), is_upi))

    # CashRecord
    for cnum in [1, 2, 3]:
        for cat, dtype, dval, count, total in cash[cnum]:
            cur.execute("""
                INSERT INTO "CashRecord"
                    (counter_id, record_date, cash_category, denomination_type,
                     denomination_value, count, total_amount)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (counter_id, record_date, cash_category, denomination_type, denomination_value)
                DO NOTHING
            """, (counter_id[cnum], sale_date, cat, dtype, dval, count, total))

    print(f"    ✓ DB seeded.")
    return True


def date_range(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--date',  help='Single date YYYY-MM-DD')
    ap.add_argument('--start', help='Start date YYYY-MM-DD')
    ap.add_argument('--end',   help='End date YYYY-MM-DD')
    ap.add_argument('--dry-run',  action='store_true')
    ap.add_argument('--no-bills', action='store_true')
    args = ap.parse_args()

    if args.date:
        dates = [date.fromisoformat(args.date)]
    elif args.start and args.end:
        dates = list(date_range(date.fromisoformat(args.start),
                                date.fromisoformat(args.end)))
    else:
        ap.error("Provide --date or --start/--end")

    conn = psycopg2.connect(**CONN)
    conn.autocommit = False
    cur  = conn.cursor()

    cur.execute('SELECT id FROM "Shop" WHERE short_name=%s', (SHOP,))
    shop_id = cur.fetchone()[0]
    cur.execute('SELECT id, master_code FROM "Product"')
    prod_id = {mc: pid for pid, mc in cur.fetchall()}
    cur.execute('SELECT id, display_order FROM "Counter" WHERE shop_id=%s ORDER BY display_order', (shop_id,))
    counter_id = {order: cid for cid, order in cur.fetchall()}
    print(f"Shop={shop_id}  Products={len(prod_id)}  Counters={counter_id}")

    # Cache workbooks
    wbs = {}
    active_dates = []

    for d in dates:
        sheet = str(d.day)
        xkey  = "I" if d.day <= 15 else "II"
        if xkey not in wbs:
            print(f"\nLoading {XLSX[xkey]} ...")
            wbs[xkey] = openpyxl.load_workbook(XLSX[xkey], data_only=True)
        wb = wbs[xkey]
        if sheet not in wb.sheetnames:
            print(f"\n  Sheet '{sheet}' not found in {xkey} — skipping {d}")
            continue

        permit_rec = PERMIT_REC.get(d.isoformat(), {})
        rows = list(wb[sheet].iter_rows(values_only=True))
        seeded = seed_day(cur, shop_id, prod_id, counter_id, d, rows,
                          permit_rec, dry_run=args.dry_run)
        if seeded and not args.dry_run:
            conn.commit()
            active_dates.append(d)

    cur.close()
    conn.close()

    if not active_dates or args.no_bills or args.dry_run:
        print(f"\nSeeded {len(active_dates)} day(s). Bills skipped.")
        return

    print(f"\n{'─'*50}\nGenerating bills for {len(active_dates)} day(s)...")
    for d in active_dates:
        r = subprocess.run(
            ['python3', GENERATE_SCRIPT, '--date', d.isoformat()],
            capture_output=True, text=True, cwd=os.path.dirname(__file__)
        )
        print(r.stdout.rstrip())
        if r.returncode != 0:
            print(f"  BILL ERROR: {r.stderr.rstrip()}")


if __name__ == '__main__':
    main()
